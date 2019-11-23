"""
     此文件下的函数，补全 openpyxl 读取 xlsx 文件时候的却边问题
"""
from openpyxl.styles import borders, colors, Border, Side
black = Side(border_style=borders.BORDER_THIN, color=colors.BLACK)        # 黑线
big_black = Side(border_style=borders.BORDER_MEDIUM, color=colors.BLACK)  # 加粗黑线

# 初始化Mode\1.xlsx
# ws 是一个sheet表 例如 ws = wb.worksheets[0]
def iniexcel1(ws):
    # 设置打印格式
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE    # 横向打印
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # 纸张大小
    ws.page_margins.left = 0                            # 0.2*2.5
    ws.page_margins.right = 0.
    ws.page_margins.top = 0
    ws.page_margins.bottom = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.print_area = 'A1:P24'
    # 设置边框
    border6 = Border(left=black, right=black, top=black, bottom=black)
    border2 = Border(left=black, right=big_black, top=big_black, bottom=black)
    border3 = Border(left=black, right=black, top=big_black, bottom=black)
    ws['F6'].border = border6
    ws['I5'].border = border3
    ws['J5'].border = border3
    ws['L5'].border = border3
    ws['N5'].border = border3
    ws['O5'].border = border3
    ws['P5'].border = border2
    return ws

# 初始化Mode\2.xlsx
# ws 是一个sheet表 例如 ws = wb.worksheets[0]
def iniexcel2(ws):
    # 设置打印格式
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # 纵向打印
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # 纸张大小
    ws.page_margins.left = 0  # 0.2*2.5
    ws.page_margins.right = 0
    ws.page_margins.top = 0
    ws.page_margins.bottom = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.print_area = 'A1:J54'
    # 设置边框
    border1 = Border(left=black, right=black, top=black, bottom=black)
    for i in range(8, 53):
        for j in range(ord('A'), ord('J')+1):
            if j == ord('E'):
                continue
            ws[str(chr(j))+str(i)].border = border1
    ws['C6'].border = border1
    ws['C7'].border = border1
    return ws

# 初始化Mode\3.xlsx
# ws 是一个sheet表 例如 ws = wb.worksheets[0]
def iniexcel3(ws):
    # 设置打印格式
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # 纵向打印
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # 纸张大小
    ws.page_margins.left = 0  # 0.2*2.5
    ws.page_margins.right = 0
    ws.page_margins.top = 0.2
    ws.page_margins.bottom = 0.2
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    ws.print_area = 'A1:G28'
    # 设置边框（无）
    return ws


def iniexcel4(ws):
    # 设置打印格式
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # 纵向打印
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # 纸张大小
    ws.page_margins.left = 0.7  # 0.2*2.5
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0
    ws.page_margins.bottom = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.print_area = 'A1:N21'
    # 设置边框
    ws['C7'].border = Border(top=big_black, right=black)
    ws['N19'].border = Border(top=black, right=big_black, bottom=big_black)
    ws['N18'].border = Border(top=black, right=big_black, bottom=black, left=black)
    ws['N17'].border = Border(top=black, right=big_black, bottom=black, left=black)
    ws['N16'].border = Border(top=black, right=big_black, bottom=black, left=black)
    ws['N15'].border = Border(top=black, right=big_black, bottom=black, left=black)
    ws['N14'].border = Border(top=black, right=big_black, bottom=black, left=black)
    for i in range(66,78):
        ws[chr(i)+'19'].border = Border(top=black, right=black, bottom=big_black, left=black)
    for i in range(14,19):
        for j in range(68,78):
            ws[chr(j) + str(i)].border = Border(top=black, right=black, bottom=black, left=black)
    return ws

def iniexcel5(ws):
    # 设置打印格式
    ws.print_options.horizontalCentered = True
    # ws.print_options.verticalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # 纵向打印
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # 纸张大小
    ws.page_margins.left = 0.7  # 0.2*2.5
    ws.page_margins.right = 0.7
    ws.page_margins.top = 1.2
    ws.page_margins.bottom = 0
    ws.page_margins.header = 1.2
    ws.page_margins.footer = 0
    ws.print_area = 'A1:I21'
    # 设置边框
    default = Side(border_style=None, color='FF000000')
    black = Side(border_style=borders.BORDER_THIN, color=colors.BLACK)
    border = Border(left=black, right=black, top=black, bottom=black)
    border1 = Border(left=black, right=black, top=default, bottom=black)
    border2 = Border(left=black, right=black, top=black, bottom=default)
    for i in range(8, 16):
        for j in range(97, 97+8):
            if chr(j) + str(i) == 'C9' or chr(j) + str(i) == 'C10':
                continue
            ws[chr(j) + str(i)].border = border
    ws['C10'].border = border1
    ws['C9'].border = border2
    ws['I10'].border = border
    ws['I12'].border = border
    ws['I15'].border = border
    return ws
def iniexcel6(ws):
    # 设置打印格式
    ws.print_options.horizontalCentered = True
    # ws.print_options.verticalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # 纵向打印
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # 纸张大小
    ws.page_margins.left = 0.8  # 0.2*2.5
    ws.page_margins.right = 0.8
    ws.page_margins.top = 1.2
    ws.page_margins.bottom = 0
    ws.page_margins.header = 1.2
    ws.page_margins.footer = 0
    ws.print_area = 'A1:F23'
    # 设置边框
    return ws
def iniexcel7(ws):
    # 设置打印格式
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE    # 横向打印
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # 纸张大小
    ws.page_margins.left = 0.7  # 0.2*2.5
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0
    ws.page_margins.bottom = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.print_area = 'A1:S30'
    # 设置边框
    default = Side(border_style=None, color='FF000000')
    big_black = Side(border_style=borders.BORDER_MEDIUM, color=colors.BLACK)
    black = Side(border_style=borders.BORDER_THIN, color=colors.BLACK)
    border1 = Border(left=black, right=black, top=big_black, bottom=black)
    border2 = Border(left=default, right=default, top=big_black, bottom=default)
    border3 = Border(left=black, right=big_black, top=big_black, bottom=black)
    border4 = Border(left=black, right=big_black, top=black, bottom=big_black)
    border5 = Border(left=black, right=big_black, top=black, bottom=black)
    border6 = Border(left=black, right=black, top=black, bottom=black)
    border7 = Border(left=black, right=black, top=black, bottom=big_black)
    border8 = Border(left=black, right=black, top=default, bottom=black)
    list1 = ['7','10','18','23']
    for i in range(99,98+17):
        for j in list1:
            ws[chr(i)+j].border = border1
    list1 = ['28']
    for i in range(98,98+17):
        for j in list1:
            ws[chr(i)+j].border = border2
    list1 = ['S7','S10','S18','S23']
    for j in list1:
        ws[j].border = border3
        list1 = ['S9', 'S17', 'S27']
    for j in list1:
        ws[j].border = border4
    list1 = ['S8', 'S10', 'S11', 'S12', 'S13', 'S14', 'S15', 'S16', 'S21', 'S22', 'S26', 'S19', 'S20', 'S24', 'S25']
    for j in list1:
        ws[j].border = border5
    list1 = ['8','11','12','13','14','15','16','19','21','24','26']
    for i in range(98,98+10):
        for j in list1:
            ws[chr(i) + j].border = border6
    ws['O12'].border = border6
    ws['H17'].border = border7;ws['H22'].border = border7;ws['K22'].border = border7;ws['H27'].border = border7
    ws['B11'].border = border6
    ws['N8'].border = border6
    ws['N12'].border = border6
    ws['N14'].border = border6
    ws['N15'].border = border6
    ws['N20'].border = border6
    ws['N21'].border = border6
    ws['N25'].border = border6
    ws['N26'].border = border6
    ws['N19'].border = border6
    ws['N24'].border = border6
    ws['P13'].border = border6
    ws['Q12'].border = border6
    ws['B11'].border = border8
    return ws
def iniexcel8(ws):
    # 设置打印格式
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # 纵向打印
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # 纸张大小
    ws.page_margins.left = 0  # 0.2*2.5
    ws.page_margins.right = 0
    ws.page_margins.top = 0
    ws.page_margins.bottom = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.print_area = 'A1:G21'
    # 设置边框（无）

def iniexcel9(ws):
    # 设置打印格式
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # 纵向打印
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # 纸张大小
    ws.page_margins.left = 0.7  # 0.2*2.5
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0
    ws.page_margins.bottom = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.print_area = 'A1:M33'
    # 设置边框
    default = Side(border_style=None, color='FF000000')
    big_black = Side(border_style=borders.BORDER_MEDIUM, color=colors.BLACK)
    black = Side(border_style=borders.BORDER_THIN, color=colors.BLACK)
    border1 = Border(left=black, right=big_black, top=black, bottom=big_black)
    border2 = Border(left=black, right=big_black, top=big_black, bottom=black)
    border3 = Border(left=black, right=black, top=big_black, bottom=black)
    border4 = Border(left=black, right=black, top=black, bottom=big_black)
    border5 = Border(left=black, right=big_black, top=black, bottom=black)
    border6 = Border(left=black, right=black, top=black, bottom=black)
    border7 = Border(left=big_black, right=black, top=black, bottom=big_black)
    border8 = Border(left=black, right=default, top=black, bottom=black)
    list1 = ['M10','M19','M22','M25']
    for i in list1:
        ws[i].border = border1
    list1 = ['M11','M20','M23','M26', 'M6', 'M30']
    for i in list1:
        ws[i].border = border2
    list1 = ['C6','D6','E6','D11','G11','H11','D20','D23','H20','H23','I20',
             'I23','C26','D26','E26','F26','G26','H26','I26','G20','G23']
    for i in list1:
        ws[i].border = border3
    list1 = ['D30','G30']
    for i in list1:
        ws[i].border = border4
    list1 = ['M8','M9','M17','M18','M21','M24', 'M7', 'M12', 'M13', 'M14', 'M27', 'M28', 'M29']
    for i in list1:
        ws[i].border = border5
    list1 = ['H8','J8','I9','J9','C9','D9','F10','I10','J10','E12','E13','E14'
        ,'G12','G13','G14','I12','I13','I14','K12','K13','D18','K18','H18','J18'
             ,'D21','H21','J21','K21','D24','H24','I24','G27','G29',
             'I18','I21','K24','D28','J26','K27','B27','C8']
    for i in list1:
        ws[i].border = border6
    ws['I11'].border = border3
    ws['A19'].border = border7
    ws['A22'].border = border7
    ws['A25'].border = border7
    ws['D8'].border = border6
    ws['E8'].border = border6
    ws['D7'].border = border6
    ws['E7'].border = border6
    ws['F10'].border = border4
    ws['M15'].border = border1
    ws['M30'].border = border1
    ws['L10'].border = border4
    ws['L19'].border = border4
    ws['L22'].border = border4
    ws['L25'].border = border4
    ws['L8'].border = border6
    ws['L18'].border = border6
    ws['L21'].border = border6
    ws['L24'].border = border6
    ws['L27'].border = border6
    ws['K14'].border = border8
    return ws

def iniexcel10(ws):
    # 设置打印格式
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # 纵向打印
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # 纸张大小
    ws.page_margins.left = 0  # 0.2*2.5
    ws.page_margins.right = 0
    ws.page_margins.top = 0
    ws.page_margins.bottom = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.print_area = 'A1:F30'
    # 设置边框
    black = Side(border_style=borders.BORDER_THIN, color=colors.BLACK)
    border = Border(left=black, right=black, top=black, bottom=black)
    for i in range (97,97+6):
        for j in range(7,28):
            ws[chr(i)+str(j)].border = border
    return ws

def iniexcel11(ws):
    # 设置打印格式
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # 纵向打印
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # 纸张大小
    ws.page_margins.left = 0  # 0.2*2.5
    ws.page_margins.right = 0
    ws.page_margins.top = 0
    ws.page_margins.bottom = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.print_area = 'A1:28'
    # 设置边框
    ws = wb[wb.sheetnames[0]]
    black = Side(border_style=borders.BORDER_THIN, color=colors.BLACK)
    border = Border(left=black, right=black, top=black, bottom=black)
    for i in range (98,97+9):
        for j in range(8,24):
            ws[chr(i)+str(j)].border = border
    return ws
def iniexcel(ws , number):
    if number == 1 :
        iniexcel1(ws)
    elif number == 2 :
        iniexcel2(ws)
    elif number == 3 :
        iniexcel3(ws)
    elif number == 4 :
        iniexcel4(ws)
    elif number == 5 :
        iniexcel5(ws)
    elif number == 6 :
        iniexcel6(ws)
    elif number == 7 :
        iniexcel7(ws)
    elif number == 8 :
        iniexcel8(ws)
    elif number == 9 :
        iniexcel9(ws)
    elif number == 10 :
        iniexcel10(ws)
    elif number == 11 :
        iniexcel11(ws)
    else :
        print('初始化表格传入了错误的number')
if __name__ == '__main__':
    import openpyxl
    wb = openpyxl.load_workbook('../Mode/8.xlsx')
    ws = wb.worksheets[0]
    iniexcel8(ws)
    ws = wb.copy_worksheet(ws)
    iniexcel8(ws)
    wb.save('test.xlsx')
