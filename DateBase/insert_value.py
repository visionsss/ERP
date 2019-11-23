# coding=utf-8
"""
Author: vision
date: 2019/4/3 18:53
"""
from DateBase.creat import Parameter, ConcreteMix, CementAttributeDatum
from DateBase.connect_db import session
from openpyxl import load_workbook
from Util.round import get_float


def parm_init():
    parm = Parameter(
        MinC_Strength=1.25,
        MaxC_Strength=1.35,
        MinS_FinenessDensity=2.3,
        MaxS_FinenessDensity=3.0,
        MinS_SurfaceDensity=2600,
        MaxS_SurfaceDensity=2800,
        MinS_Density=1430,
        MaxS_Density=1550,
        MinS_SlitContent=0.8,
        MaxS_SlitContent=0.8,
        MinS_WaterContent=0.05,
        MaxS_WaterContent=0.07,
        MinG_GrainContent=4.5,
        MaxG_GrainContent=5,
        MinG_CrushLevel=6,
        MaxG_CrushLevel=8,
        MinG_Density=1430,
        MaxG_Density=1550,
        MinG_SlitContent=0.9,
        MaxG_SlitContent=0.9,
        MinG_WaterContent=0.9,
        MaxG_WaterContent=1.6,
        MinA_Density=10,
        MaxA_Density=12,
        MinR7_Compression=0.8,
        MaxR7_Compression=1,
        MinR28_Compression=1.2,
        MaxR28_Compression=1.5,
        Project1Manager='a留空',                                 # 表1工程负责人
        Project1FillSheeter='梁华君',                             # 表1填表人
        Project2InspectCodeEdit='GD-C3-523□□□',                        # 表2检验码
        Project2Manager='杜敏燕',                                 # 表2单位检验工程负责人
        Project2Checker='毛湘平',                                # 表2校核
        Project2Try='梁建南',                                     # 表2检验
        Project3MakeSheet='梁华君',                               # 表3制表
        Project3InspectCodeEdit='GD-C1-331□□□',                        # 表3检验码
        Project4Manager='杜敏燕',                                 # 表4单位工程负责人
        Project4Checker='毛湘平',                                 # 表4复核
        Project4Calculate='梁华君',                               # 表4计算
        Project4InspectCodeEdit='GD-C1-334□□',                        # 表4检验码
        Project5Manager='杜敏燕',  # 表5技术负责人
        Project5Filler='梁华君',            # 表5技术填表人
        Project7ConDesignSpeciEdit='JGJ55-2011',                 # 表7实验规程：
        Project7CodeEdit='GD-C3-521□□□',  # 表7省统表编号
        Project7Manager='杜敏燕',                # 表7技术负责人
        Project7Checker='毛湘平',               # 表7校核
        Project7try='高祖华',              # 表7试验
        Project9Manager='a留空',            # 表9工程负责人
        Project9Checker='卢宪文',            # 表9质检员
        Project9Record='李佳良',           # 表9记录
        Project10ConTestReportTestBasisEdit='GB/T 50082-2009',           # 表10检验依据
        Project10InspectCodeEdit='GD-C3-527□□□',          # 表10检验码
        Project10TimeEdit='9:30',  # 加水压日期的时分
        Project10MaxCreepEdit=130,     # 最高渗水值最大区间
        Project10MinCreepEdit=1,  # 最低渗水值最小区间
        Project10Manager='杜敏燕',   # 表10技术负责人
        Project10Examine='毛湘平',   # 表10审核
        Project10Checker='李杰和',   # 表10检验人
        Project11Manager='a留空',   # 表11技术负责人
        Project11Checker='a留空',    # 表11试验
    )
    session.add(parm)
    session.commit()


def insert_concrete_mix(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    if ws['A2'].value != '配合比选用汇总表':
        raise Exception('表格选错了')
    rows = ws.max_row
    session.execute('delete from concrete_mix')
    for i in range(5, rows + 1):
        ConcreteName = ws['A' + str(i)].value
        if ConcreteName is None:
            break
        MixRatioID = ws['B' + str(i)].value
        StrengthLevel = ws['D' + str(i)].value
        ImperLevel = ws['E' + str(i)].value
        if ImperLevel == '/':
            ImperLevel = None
        SwellLevel = ws['F' + str(i)].value
        if SwellLevel == '/':
            SwellLevel = None
        MixRatioName = ws['G' + str(i)].value
        SlumpNum = str(ws['H' + str(i)].value).replace(' ', '')
        StandardDeviation = ws['I' + str(i)].value
        ConcreteStrengh = get_float(ws['J' + str(i)].value, 1)
        WaterNum = ws['K' + str(i)].value
        CementNum = ws['L' + str(i)].value
        FlyashNum = ws['M' + str(i)].value
        SandNum = ws['N' + str(i)].value
        GravelNum = ws['O' + str(i)].value
        CementRatio = get_float(ws['P' + str(i)].value, 2)
        SandRatio = get_float(ws['Q' + str(i)].value, 3)
        AdmixtureAmount = ws['R' + str(i)].value
        AdmixtureNum = get_float(ws['S' + str(i)].value, 1)
        SwellingNum = get_float(ws['T' + str(i)].value, 0)
        # print(i , ws['U' + str(i)].value)
        MassDensity = get_float(ws['U' + str(i)].value, 0)
        InitialTime = ws['V' + str(i)].value
        FinalTime = ws['W' + str(i)].value
        mix = ConcreteMix(
            ConcreteName=ConcreteName,
            MixRatioID=MixRatioID,
            StrengthLevel=StrengthLevel,
            ImperLevel=ImperLevel,
            SwellLevel=SwellLevel,
            MixRatioName=MixRatioName,
            SlumpNum=SlumpNum,
            StandardDeviation=StandardDeviation,
            ConcreteStrengh=ConcreteStrengh,
            WaterNum=WaterNum,
            CementNum=CementNum,
            FlyashNum=FlyashNum,
            SandNum=SandNum,
            GravelNum=GravelNum,
            CementRatio=CementRatio,
            SandRatio=SandRatio,
            AdmixtureAmount=AdmixtureAmount,
            AdmixtureNum=AdmixtureNum,
            SwellingNum=SwellingNum,
            MassDensity=MassDensity,
            InitialTime=InitialTime,
            FinalTime=FinalTime,
        )
        session.add(mix)
        session.commit()


def insert_cement_attribute_data(path):
    wb = load_workbook(path)
    ws = wb.worksheets[0]
    if ws['A1'].value != '水泥资料一览表':
        raise Exception('表格选错了')
    rows = ws.max_row
    session.execute('delete from cement_attribute_data')
    for i in range(3, rows + 1):
        ArrivalTime = ws['A' + str(i)].value
        if ArrivalTime is None:
            break
        CementVariety = ws['B' + str(i)].value
        Manufacturer = ws['C' + str(i)].value
        ProductionDate = ws['D' + str(i)].value
        CementId = ws['E' + str(i)].value
        CementNumber = ws['F' + str(i)].value
        IsStability = ws['G' + str(i)].value
        if IsStability == '合格':
            IsStability = 1
        elif IsStability == '不合格':
            IsStability = 0
        else:
            IsStability = 1
        InitialTime = str(ws['H' + str(i)].value).replace(':00', '')
        FinalTime = str(ws['I' + str(i)].value).replace(':00', '')
        R3_Compression = ws['J' + str(i)].value
        R28_Compression = ws['K' + str(i)].value
        R3_Bending = ws['L' + str(i)].value
        R28_Bending = ws['M' + str(i)].value
        PriorityLevel = ws['N' + str(i)].value
        atr = CementAttributeDatum(
            ArrivalTime=ArrivalTime,
            CementVariety=CementVariety,
            Manufacturer=Manufacturer,
            ProductionDate=ProductionDate,
            CementId=CementId,
            CementNumber=CementNumber,
            IsStability=IsStability,
            InitialTime=InitialTime,
            FinalTime=FinalTime,
            R3_Compression=R3_Compression,
            R28_Compression=R28_Compression,
            R3_Bending=R3_Bending,
            R28_Bending=R28_Bending,
            PriorityLevel=PriorityLevel)
        session.add(atr)
        session.commit()


if __name__ == "__main__":
    parm_init()
    mix_path = '../配合比选用汇总表.xlsx'
    insert_concrete_mix(mix_path)
    attribute_path = '../水泥一览表.xlsx'
    insert_cement_attribute_data(attribute_path)
