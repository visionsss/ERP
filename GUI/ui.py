# coding=utf-8
"""
Author: vision
date: 2019/4/4 19:51
"""
from GUI.ui_mode import Ui_Form
from Util.get_parm import get_parm
from PyQt5 import QtCore
from Util.PictureUtil.GetPicName import getpiname
from PyQt5.QtSql import (QSqlDatabase, QSqlTableModel)
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt
from DateBase.connect_db import db_path
from DateBase.creat import Parameter
from DateBase.connect_db import session
from DateBase.creat import CementAttributeDatum, ConcreteMix
from openpyxl import Workbook
from Excel.CreateExcel.CreateConUseBuyRecord17 import ConUseBuyRecord
from Excel.CreateExcel.CreateConStrengReport234 import ConStrengReport
from Excel.CreateExcel.CreateConFactoryCard56 import CreateConFacyoryCard, CreateSlumpCard
from Excel.CreateExcel.CreateConUseProve8 import CreteConUseProve
from Excel.CreateExcel.CreateConQualityRecord9 import ConQualityRecord
from Excel.CreateExcel.CreatePermeabilityTestReport10 import PermeabilityTestReport
from openpyxl import load_workbook
from Util.history_Util.history import his


class Ui(Ui_Form):
    def function(self):
        self.OutExcelButton.clicked.connect(self.OutExcel)  # 导出表格按钮
        self.InSQLButton.clicked.connect(self.InSQL)  # Excel导入信息按钮
        self.OutPutSQLButton.clicked.connect(self.OutSQL)      # 基本信息导出按钮
        self.CementDateButton.clicked.connect(self.CementDate)  # 水泥一览表信息按钮
        self.MixpushButton.clicked.connect(self.MixDate)  # 配合比表信息按钮
        self.ChangePrarmerButton.clicked.connect(
            self.ChangePrarmer)  # 确认修改参数按钮
        self.CementrefreshButton.clicked.connect(self.cement_refresh)  # 刷新
        self.CementsubmitButton.clicked.connect(self.cement_submit)  # 提交
        self.CementdelButton.clicked.connect(self.cement_del)  # 删除
        self.CementQueryButton.clicked.connect(self.query_cement)  # 查询
        self.mixrefreshButton.clicked.connect(self.mix_refresh)  # 刷新
        self.mixsubmitButton.clicked.connect(self.mix_submit)  # 提交
        self.MixDelButton.clicked.connect(self.mix_del)  # 删除
        self.MixqueryButton.clicked.connect(self.query_mix)  # 查询
        self.deletArtButton.clicked.connect(self.empty_Art)  # 水泥一览表数据清空
        self.deleMIXButton.clicked.connect(self.empty_Mix)  # 配合比表数据清空
        self.CementAddButton.clicked.connect(self.addArt)   # 添加水泥按钮
        self.mixAddButton.clicked.connect(self.addMix)      # 添加配合比按钮
        self.CementInfoButton.clicked.connect(self.CementInfo)  # 选择水泥一览表记录
        self.MixPoportionButton.clicked.connect(self.ConMixInsert)  # 选择配合比表记录
        self.ChoicConcreteUsageRecordButton.clicked.connect(
            self.ChoicConcreteUsageRecord)  # 选择使用记录表
        self.OutSQLButton_2.clicked.connect(self.ChoicSQLPath)  # 选择OutPutSQL路径
        self.OutSQLButton_3.clicked.connect(self.OutPutSQL)  # 导出SQL
        self.OutPutButton.clicked.connect(self.ouputexcel)   # 确认导出按钮
        self.OutPutVoidButton.clicked.connect(self.outputvoidexcel)  # 确认导出按钮
        # 查看数据库是否存在
        try:
            parm = get_parm()
            print('已获取SQL')
        except BaseException:
            print('创建SQL')
            from DateBase.creat import creat_table
            from DateBase.insert_value import parm_init
            creat_table()
            parm_init()
        # 填充信息
        parm = get_parm()
        self.MinC_StrengthEdit.setText(str(parm.MinC_Strength))
        self.MaxC_StrengthEdit.setText(str(parm.MaxC_Strength))
        self.MinS_FinenessDensityEdit.setText(str(parm.MinS_FinenessDensity))
        self.MaxS_FinenessDensityEdit.setText(str(parm.MaxS_FinenessDensity))
        self.MinS_SurfaceDensityEdit.setText(str(parm.MinS_SurfaceDensity))
        self.MaxS_SurfaceDensityEdit.setText(str(parm.MaxS_SurfaceDensity))
        self.MinS_DensityEdit.setText(str(parm.MinS_Density))
        self.MaxS_DensityEdit.setText(str(parm.MaxS_Density))
        self.MinS_SlitContentEdit.setText(str(parm.MinS_SlitContent))
        self.MaxS_SlitContentEdit.setText(str(parm.MaxS_SlitContent))
        self.MinS_WaterContentEdit.setText(str(parm.MinS_WaterContent))
        self.MaxS_WaterContentEdit.setText(str(parm.MaxS_WaterContent))
        self.MinG_GrainContentEdit.setText(str(parm.MinG_GrainContent))
        self.MaxG_GrainContentEdit.setText(str(parm.MaxG_GrainContent))
        self.MinG_CrushLevelEdit.setText(str(parm.MinG_CrushLevel))
        self.MaxG_CrushLevelEdit.setText(str(parm.MaxG_CrushLevel))
        self.MinG_DensityEdit.setText(str(parm.MinG_Density))
        self.MaxG_DensityEdit.setText(str(parm.MaxG_Density))
        self.MinG_SlitContentEdit.setText(str(parm.MinG_SlitContent))
        self.MaxG_SlitContentEdit.setText(str(parm.MaxG_SlitContent))
        self.MinG_WaterContentEdit.setText(str(parm.MinG_WaterContent))
        self.MaxG_WaterContentEdit.setText(str(parm.MaxG_WaterContent))
        self.MinA_DensityEdit.setText(str(parm.MinA_Density))
        self.MaxA_DensityEdit.setText(str(parm.MaxA_Density))
        self.MinR7_CompressionEdit.setText(str(parm.MinR7_Compression))
        self.MaxR7_CompressionEdit.setText(str(parm.MaxR7_Compression))
        self.MinR28_CompressionEdit.setText(str(parm.MinR28_Compression))
        self.MaxR28_CompressionEdit.setText(str(parm.MaxR28_Compression))
        _translate = QtCore.QCoreApplication.translate
        PicName = getpiname()
        for i in range(len(PicName)):
            self.Project1Manager.addItem("")
            self.Project1Manager.setItemText(i, _translate("Form", PicName[i]))
            self.Project1FillSheeter.addItem("")
            self.Project1FillSheeter.setItemText(
                i, _translate("Form", PicName[i]))
            self.Project2Manager.addItem("")
            self.Project2Manager.setItemText(i, _translate("Form", PicName[i]))
            self.Project2Checker.addItem("")
            self.Project2Checker.setItemText(i, _translate("Form", PicName[i]))
            self.Project2Try.addItem("")
            self.Project2Try.setItemText(i, _translate("Form", PicName[i]))
            self.Project3MakeSheet.addItem("")
            self.Project3MakeSheet.setItemText(
                i, _translate("Form", PicName[i]))
            self.Project4Manager.addItem("")
            self.Project4Manager.setItemText(i, _translate("Form", PicName[i]))
            self.Project4Checker.addItem("")
            self.Project4Checker.setItemText(i, _translate("Form", PicName[i]))
            self.Project4Calculate.addItem("")
            self.Project4Calculate.setItemText(
                i, _translate("Form", PicName[i]))
            self.Project5Manager.addItem("")
            self.Project5Manager.setItemText(i, _translate("Form", PicName[i]))
            self.Project5Filler.addItem("")
            self.Project5Filler.setItemText(i, _translate("Form", PicName[i]))
            self.Project7Manager.addItem("")
            self.Project7Manager.setItemText(i, _translate("Form", PicName[i]))
            self.Project7Checker.addItem("")
            self.Project7Checker.setItemText(i, _translate("Form", PicName[i]))
            self.Project7try.addItem("")
            self.Project7try.setItemText(i, _translate("Form", PicName[i]))
            self.Project9Manager.addItem("")
            self.Project9Manager.setItemText(i, _translate("Form", PicName[i]))
            self.Project9Checker.addItem("")
            self.Project9Checker.setItemText(i, _translate("Form", PicName[i]))
            self.Project9Record.addItem("")
            self.Project9Record.setItemText(i, _translate("Form", PicName[i]))
            self.Project10Manager.addItem("")
            self.Project10Manager.setItemText(
                i, _translate("Form", PicName[i]))
            self.Project10Examine.addItem("")
            self.Project10Examine.setItemText(
                i, _translate("Form", PicName[i]))
            self.Project10Checker.addItem("")
            self.Project10Checker.setItemText(
                i, _translate("Form", PicName[i]))
            self.Project11Manager.addItem("")
            self.Project11Manager.setItemText(
                i, _translate("Form", PicName[i]))
            self.Project11Checker.addItem("")
            self.Project11Checker.setItemText(
                i, _translate("Form", PicName[i]))
        self.Project1Manager.setCurrentText(str(parm.Project1Manager))
        self.Project1FillSheeter.setCurrentText(str(parm.Project1FillSheeter))
        self.Project2InspectCodeEdit.setText(str(parm.Project2InspectCodeEdit))
        self.Project2Manager.setCurrentText(str(parm.Project2Manager))
        self.Project2Checker.setCurrentText(str(parm.Project2Checker))
        self.Project2Try.setCurrentText(str(parm.Project2Try))
        self.Project3MakeSheet.setCurrentText(str(parm.Project3MakeSheet))
        self.Project3InspectCodeEdit.setText(str(parm.Project3InspectCodeEdit))
        self.Project4Manager.setCurrentText(str(parm.Project4Manager))
        self.Project4Checker.setCurrentText(str(parm.Project4Checker))
        self.Project4Calculate.setCurrentText(str(parm.Project4Calculate))
        self.Project4InspectCodeEdit.setText(str(parm.Project4InspectCodeEdit))
        self.Project5Manager.setCurrentText(str(parm.Project5Manager))
        self.Project5Filler.setCurrentText(str(parm.Project5Filler))
        self.Project7ConDesignSpeciEdit.setText(
            str(parm.Project7ConDesignSpeciEdit))
        self.Project7CodeEdit.setText(str(parm.Project7CodeEdit))
        self.Project7Manager.setCurrentText(str(parm.Project7Manager))
        self.Project7Checker.setCurrentText(str(parm.Project7Checker))
        self.Project7try.setCurrentText(str(parm.Project7try))
        self.Project9Manager.setCurrentText(str(parm.Project9Manager))
        self.Project9Checker.setCurrentText(str(parm.Project9Checker))
        self.Project9Record.setCurrentText(str(parm.Project9Record))
        self.Project10ConTestReportTestBasisEdit.setText(
            str(parm.Project10ConTestReportTestBasisEdit))
        self.Project10InspectCodeEdit.setText(
            str(parm.Project10InspectCodeEdit))
        from PyQt5.QtCore import QTime
        time = str(parm.Project10TimeEdit).split(':')
        time = QTime(int(time[0]), int(time[1]))
        self.Project10TimeEdit.setTime(time)
        self.Project10MaxCreepEdit.setText(str(parm.Project10MaxCreepEdit))
        self.Project10MinCreepEdit.setText(str(parm.Project10MinCreepEdit))
        self.Project10Manager.setCurrentText(str(parm.Project10Manager))
        self.Project10Examine.setCurrentText(str(parm.Project10Examine))
        self.Project10Checker.setCurrentText(str(parm.Project10Checker))
        self.Project11Manager.setCurrentText(str(parm.Project11Manager))
        self.Project11Checker.setCurrentText(str(parm.Project11Checker))
        # 插入水泥一览表
        self.con2 = QSqlDatabase.addDatabase('QSQLITE')
        self.con2.setDatabaseName(db_path)
        # con2.exec_("PRAGMA foreign_keys = ON;")

        self.CementMode = QSqlTableModel()
        self.CementMode.setTable("cement_attribute_data")
        self.CementMode.setSort(0, Qt.AscendingOrder)
        self.CementMode.setEditStrategy(self.CementMode.OnManualSubmit)
        self.CementMode.setHeaderData(0, Qt.Horizontal, "id")
        self.CementMode.setHeaderData(1, Qt.Horizontal, "进场日期")
        self.CementMode.setHeaderData(2, Qt.Horizontal, "水泥品种")
        self.CementMode.setHeaderData(3, Qt.Horizontal, "生产厂家")
        self.CementMode.setHeaderData(4, Qt.Horizontal, "生产日期")
        self.CementMode.setHeaderData(5, Qt.Horizontal, "编号")
        self.CementMode.setHeaderData(6, Qt.Horizontal, "数量(T)")
        self.CementMode.setHeaderData(7, Qt.Horizontal, "安定性")
        self.CementMode.setHeaderData(8, Qt.Horizontal, "初凝")
        self.CementMode.setHeaderData(9, Qt.Horizontal, "终凝")
        self.CementMode.setHeaderData(10, Qt.Horizontal, "R3抗压")
        self.CementMode.setHeaderData(11, Qt.Horizontal, "R28抗压")
        self.CementMode.setHeaderData(12, Qt.Horizontal, "R3抗折")
        self.CementMode.setHeaderData(13, Qt.Horizontal, "R28抗折")
        self.CementMode.setHeaderData(14, Qt.Horizontal, "是否优先")
        self.CementMode.select()
        self.CementtableView.setModel(self.CementMode)
        self.CementtableView.setSelectionMode(QTableView.SingleSelection)
        self.CementtableView.setSelectionBehavior(QTableView.SelectRows)
        self.CementtableView.resizeColumnsToContents()
        # 配合比表
        self.MixMode = QSqlTableModel()
        self.MixMode.setTable("concrete_mix")
        self.MixMode.setSort(0, Qt.AscendingOrder)
        self.MixMode.setEditStrategy(QSqlTableModel.OnManualSubmit)
        self.MixMode.setHeaderData(1 - 1, Qt.Horizontal, "名称")
        self.MixMode.setHeaderData(2 - 1, Qt.Horizontal, "配合比编号")
        self.MixMode.setHeaderData(3 - 1, Qt.Horizontal, "强度等级 ")
        self.MixMode.setHeaderData(4 - 1, Qt.Horizontal, "抗渗等级")
        self.MixMode.setHeaderData(5 - 1, Qt.Horizontal, "膨胀")
        self.MixMode.setHeaderData(6 - 1, Qt.Horizontal, "配合比编号2")
        self.MixMode.setHeaderData(7 - 1, Qt.Horizontal, "坍落度")
        self.MixMode.setHeaderData(8 - 1, Qt.Horizontal, "标准差(MPa)")
        self.MixMode.setHeaderData(9 - 1, Qt.Horizontal, "配制强度(MPa)")
        self.MixMode.setHeaderData(10 - 1, Qt.Horizontal, "水W")
        self.MixMode.setHeaderData(11 - 1, Qt.Horizontal, "水泥C")
        self.MixMode.setHeaderData(12 - 1, Qt.Horizontal, "粉煤灰F")
        self.MixMode.setHeaderData(13 - 1, Qt.Horizontal, "砂S")
        self.MixMode.setHeaderData(14 - 1, Qt.Horizontal, "石G")
        self.MixMode.setHeaderData(15 - 1, Qt.Horizontal, "水胶比A/P")
        self.MixMode.setHeaderData(16 - 1, Qt.Horizontal, "砂率 BS")
        self.MixMode.setHeaderData(17 - 1, Qt.Horizontal, "外加剂掺量A%")
        self.MixMode.setHeaderData(18 - 1, Qt.Horizontal, "外加剂用量LS-JS(B)")
        self.MixMode.setHeaderData(19 - 1, Qt.Horizontal, "膨胀剂用量")
        self.MixMode.setHeaderData(20 - 1, Qt.Horizontal, "质量密度  (容重）Mcp")
        self.MixMode.setHeaderData(21 - 1, Qt.Horizontal, "初凝时间")
        self.MixMode.setHeaderData(22 - 1, Qt.Horizontal, "终凝时间")
        self.MixMode.select()
        self.MixtableView.setModel(self.MixMode)
        self.MixtableView.setSelectionMode(QTableView.SingleSelection)
        self.MixtableView.setSelectionBehavior(QTableView.SelectRows)
        self.MixtableView.resizeColumnsToContents()
        if self.con2.isOpen():
            self.con2.close()

    def OutExcel(self):
        if self.con2.isOpen():
            self.con2.close()
        self.stackedWidget.setCurrentIndex(0)

    def InSQL(self):
        if self.con2.isOpen():
            self.con2.close()
        self.stackedWidget.setCurrentIndex(1)

    def OutSQL(self):
        if self.con2.isOpen():
            self.con2.close()
        self.stackedWidget.setCurrentIndex(2)

    def CementDate(self):
        self.stackedWidget.setCurrentIndex(3)
        self.cement_refresh()
        if not self.con2.isOpen():
            self.con2.open()
            self.CementMode = QSqlTableModel()
            self.CementMode.setTable("cement_attribute_data")
            self.CementMode.setSort(0, Qt.AscendingOrder)
            self.CementMode.setEditStrategy(self.CementMode.OnManualSubmit)
            self.CementMode.setHeaderData(0, Qt.Horizontal, "id")
            self.CementMode.setHeaderData(1, Qt.Horizontal, "进场日期")
            self.CementMode.setHeaderData(2, Qt.Horizontal, "水泥品种")
            self.CementMode.setHeaderData(3, Qt.Horizontal, "生产厂家")
            self.CementMode.setHeaderData(4, Qt.Horizontal, "生产日期")
            self.CementMode.setHeaderData(5, Qt.Horizontal, "编号")
            self.CementMode.setHeaderData(6, Qt.Horizontal, "数量(T)")
            self.CementMode.setHeaderData(7, Qt.Horizontal, "安定性")
            self.CementMode.setHeaderData(8, Qt.Horizontal, "初凝")
            self.CementMode.setHeaderData(9, Qt.Horizontal, "终凝")
            self.CementMode.setHeaderData(10, Qt.Horizontal, "R3抗压")
            self.CementMode.setHeaderData(11, Qt.Horizontal, "R28抗压")
            self.CementMode.setHeaderData(12, Qt.Horizontal, "R3抗折")
            self.CementMode.setHeaderData(13, Qt.Horizontal, "R28抗折")
            self.CementMode.setHeaderData(14, Qt.Horizontal, "是否优先")
            self.CementMode.select()
            self.CementtableView.setModel(self.CementMode)
            self.CementtableView.setSelectionMode(QTableView.SingleSelection)
            self.CementtableView.setSelectionBehavior(QTableView.SelectRows)
            self.CementtableView.resizeColumnsToContents()

    def MixDate(self):
        self.stackedWidget.setCurrentIndex(4)
        self.mix_refresh()
        if not self.con2.isOpen():
            self.con2.open()
            self.MixMode = QSqlTableModel()
            self.MixMode.setTable("concrete_mix")
            self.MixMode.setSort(0, Qt.AscendingOrder)
            self.MixMode.setEditStrategy(QSqlTableModel.OnManualSubmit)
            self.MixMode.setHeaderData(1 - 1, Qt.Horizontal, "名称")
            self.MixMode.setHeaderData(2 - 1, Qt.Horizontal, "配合比编号")
            self.MixMode.setHeaderData(3 - 1, Qt.Horizontal, "强度等级 ")
            self.MixMode.setHeaderData(4 - 1, Qt.Horizontal, "抗渗等级")
            self.MixMode.setHeaderData(5 - 1, Qt.Horizontal, "膨胀")
            self.MixMode.setHeaderData(6 - 1, Qt.Horizontal, "配合比编号2")
            self.MixMode.setHeaderData(7 - 1, Qt.Horizontal, "坍落度")
            self.MixMode.setHeaderData(8 - 1, Qt.Horizontal, "标准差(MPa)")
            self.MixMode.setHeaderData(9 - 1, Qt.Horizontal, "配制强度(MPa)")
            self.MixMode.setHeaderData(10 - 1, Qt.Horizontal, "水W")
            self.MixMode.setHeaderData(11 - 1, Qt.Horizontal, "水泥C")
            self.MixMode.setHeaderData(12 - 1, Qt.Horizontal, "粉煤灰F")
            self.MixMode.setHeaderData(13 - 1, Qt.Horizontal, "砂S")
            self.MixMode.setHeaderData(14 - 1, Qt.Horizontal, "石G")
            self.MixMode.setHeaderData(15 - 1, Qt.Horizontal, "水胶比A/P")
            self.MixMode.setHeaderData(16 - 1, Qt.Horizontal, "砂率 BS")
            self.MixMode.setHeaderData(17 - 1, Qt.Horizontal, "外加剂掺量A%")
            self.MixMode.setHeaderData(18 - 1, Qt.Horizontal, "外加剂用量LS-JS(B)")
            self.MixMode.setHeaderData(19 - 1, Qt.Horizontal, "膨胀剂用量")
            self.MixMode.setHeaderData(20 - 1, Qt.Horizontal, "质量密度  (容重）Mcp")
            self.MixMode.setHeaderData(21 - 1, Qt.Horizontal, "初凝时间")
            self.MixMode.setHeaderData(22 - 1, Qt.Horizontal, "终凝时间")
            self.MixMode.select()
            self.MixtableView.setModel(self.MixMode)
            self.MixtableView.setSelectionMode(QTableView.SingleSelection)
            self.MixtableView.setSelectionBehavior(QTableView.SelectRows)
            self.MixtableView.resizeColumnsToContents()

    def ChangePrarmer(self):
        if QMessageBox.question(
            QWidget(),
            "Question",
            "是否确定修改参数？",
            QMessageBox.Ok | QMessageBox.Cancel,
                QMessageBox.Ok) == QMessageBox.Ok:
            try:
                parm = Parameter(
                    MinC_Strength=float(self.MinC_StrengthEdit.text()),
                    MaxC_Strength=float(self.MaxC_StrengthEdit.text()),
                    MinS_FinenessDensity=float(self.MinS_FinenessDensityEdit.text()),
                    MaxS_FinenessDensity=float(self.MaxS_FinenessDensityEdit.text()),
                    MinS_SurfaceDensity=float(self.MinS_SurfaceDensityEdit.text()),
                    MaxS_SurfaceDensity=float(self.MaxS_SurfaceDensityEdit.text()),
                    MinS_Density=float(self.MinS_DensityEdit.text()),
                    MaxS_Density=float(self.MaxS_DensityEdit.text()),
                    MinS_SlitContent=float(self.MinS_SlitContentEdit.text()),
                    MaxS_SlitContent=float(self.MaxS_SlitContentEdit.text()),
                    MinS_WaterContent=float(self.MinS_WaterContentEdit.text()),
                    MaxS_WaterContent=float(self.MaxS_WaterContentEdit.text()),
                    MinG_GrainContent=float(self.MinG_GrainContentEdit.text()),
                    MaxG_GrainContent=float(self.MaxG_GrainContentEdit.text()),
                    MinG_CrushLevel=float(self.MinG_CrushLevelEdit.text()),
                    MaxG_CrushLevel=float(self.MaxG_CrushLevelEdit.text()),
                    MinG_Density=float(self.MinG_DensityEdit.text()),
                    MaxG_Density=float(self.MaxG_DensityEdit.text()),
                    MinG_SlitContent=float(self.MinG_SlitContentEdit.text()),
                    MaxG_SlitContent=float(self.MaxG_SlitContentEdit.text()),
                    MinG_WaterContent=float(self.MinG_WaterContentEdit.text()),
                    MaxG_WaterContent=float(self.MaxG_WaterContentEdit.text()),
                    MinA_Density=float(self.MinA_DensityEdit.text()),
                    MaxA_Density=float(self.MaxA_DensityEdit.text()),
                    MinR7_Compression=float(self.MinR7_CompressionEdit.text()),
                    MaxR7_Compression=float(self.MaxR7_CompressionEdit.text()),
                    MinR28_Compression=float(self.MinR28_CompressionEdit.text()),
                    MaxR28_Compression=float(self.MaxR28_CompressionEdit.text()),
                    Project1Manager=self.Project1Manager.currentText(),
                    Project1FillSheeter=self.Project1FillSheeter.currentText(),
                    Project2InspectCodeEdit=self.Project2InspectCodeEdit.text(),
                    Project2Manager=self.Project2Manager.currentText(),
                    Project2Checker=self.Project2Checker.currentText(),
                    Project2Try=self.Project2Try.currentText(),
                    Project3MakeSheet=self.Project3MakeSheet.currentText(),
                    Project3InspectCodeEdit=self.Project3InspectCodeEdit.text(),
                    Project4Manager=self.Project4Manager.currentText(),
                    Project4Checker=self.Project4Checker.currentText(),
                    Project4Calculate=self.Project4Calculate.currentText(),
                    Project4InspectCodeEdit=self.Project4InspectCodeEdit.text(),
                    Project5Manager=self.Project5Manager.currentText(),
                    Project5Filler=self.Project5Filler.currentText(),
                    Project7ConDesignSpeciEdit=self.Project7ConDesignSpeciEdit.text(),
                    Project7CodeEdit=self.Project7CodeEdit.text(),
                    Project7Manager=self.Project7Manager.currentText(),
                    Project7Checker=self.Project7Checker.currentText(),
                    Project7try=self.Project7try.currentText(),
                    Project9Manager=self.Project9Manager.currentText(),
                    Project9Checker=self.Project9Checker.currentText(),
                    Project9Record=self.Project9Record.currentText(),
                    Project10ConTestReportTestBasisEdit=self.Project10ConTestReportTestBasisEdit.text(),
                    Project10InspectCodeEdit=self.Project10InspectCodeEdit.text(),
                    Project10TimeEdit=self.Project10TimeEdit.text(),
                    Project10MaxCreepEdit=self.Project10MaxCreepEdit.text(),
                    Project10MinCreepEdit=self.Project10MinCreepEdit.text(),
                    Project10Manager=self.Project10Manager.currentText(),
                    Project10Examine=self.Project10Examine.currentText(),
                    Project10Checker=self.Project10Checker.currentText(),
                    Project11Manager=self.Project11Manager.currentText(),
                    Project11Checker=self.Project11Checker.currentText())
                session.add(parm)
                session.commit()
                QMessageBox.information(QWidget(), "修改", "成功")
            except BaseException:
                QMessageBox.information(
                    QWidget(), "错误", "修改失败！！！！\n请核对数据是否有误或留空白。")

    def cement_refresh(self):
        self.CementMode.setFilter("1=1")
        self.CementMode.select()

    def cement_submit(self):
        if (QMessageBox.question(QWidget(), "修改", "是否确定修改",
                                 QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes):
            try:
                self.CementMode.submitAll()
                QMessageBox.about(QWidget(), '修改', '修改成功')
            except BaseException:
                QMessageBox.about(QWidget(), '失败', '失败')

    def cement_del(self):
        try:
            index = self.CementtableView.currentIndex()
            self.CementMode.removeRow(index.row())
            self.CementMode.submitAll()
            self.CementMode.select()

        except BaseException:
            QMessageBox.about(QWidget(), '失败', '失败')

    def query_cement(self):
        CementId = self.CementIdEdit_2.text()
        IsPriority = self.IsPriorityEdit.text()
        sql = "1=1 "
        if CementId != '':
            sql = sql + \
                "and CementId like '%CementId_1%' ".replace('CementId_1', CementId)
        if IsPriority != '':
            sql = sql + \
                "and PriorityLevel like '%PriorityLevel_1%' ".replace('PriorityLevel_1', IsPriority)
        print(sql)
        self.CementMode.setFilter(sql)
        self.CementMode.select()

    def mix_refresh(self):
        self.MixMode.setFilter("1=1")
        self.CementMode.select()

    def mix_submit(self):
        if (QMessageBox.question(QWidget(), "修改", "是否确定修改",
                                 QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes):
            try:
                self.MixMode.submitAll()
                QMessageBox.about(QWidget(), '修改', '修改成功')
            except BaseException:
                QMessageBox.about(QWidget(), '失败', '失败')

    def mix_del(self):
        try:
            index = self.MixtableView.currentIndex()
            self.MixMode.removeRow(index.row())
            self.MixMode.submitAll()
            self.MixMode.select()

        except BaseException:
            QMessageBox.about(QWidget(), '失败', '失败')

    def query_mix(self):
        MixName = self.MixNameEdit.text()
        MixId = self.MixIdEdit.text()
        MixStrength = self.MixStrengthEdit.text()
        MixLevel = self.MixLevelEdit.text()
        sql = "1=1 "
        if MixName != '':
            sql = sql + \
                "and ConcreteName like '%MixName_1%' ".replace('MixName_1', MixName)
        if MixId != '':
            sql = sql + \
                "and MixRatioID like '%MixId_1%' ".replace('MixId_1', MixId)
        if MixStrength != '':
            sql = sql + \
                "and StrengthLevel like '%MixStrength_1%' ".replace('MixStrength_1', MixStrength)
        if MixLevel != '':
            sql = sql + \
                "and ImperLevel like '%MixLevel_1%' ".replace('MixLevel_1', MixLevel)
        print(sql)
        self.MixMode.setFilter(sql)
        self.MixMode.select()

    def empty_Art(self):
        if (QMessageBox.question(QWidget(), "删除", "是否确定删除所有数据",
                                 QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes):
            try:
                self.con2.exec_("delete from cement_attribute_data")
                self.CementMode.select()
                QMessageBox.about(QWidget(), '删除', '删除成功')
            except BaseException:
                QMessageBox.about(QWidget(), '失败', '失败')

    def empty_Mix(self):
        if (QMessageBox.question(QWidget(), "删除", "是否确定删除所有数据",
                                 QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes):
            try:
                self.con2.exec_("delete from concrete_mix")
                self.MixMode.select()
                QMessageBox.about(QWidget(), '删除', '删除成功')
            except BaseException:
                QMessageBox.about(QWidget(), '失败', '失败')

    def addArt(self):
        try:
            row = self.CementMode.rowCount()
            self.CementMode.insertRow(row)
            index = self.CementMode.index(row, 0)
            self.CementtableView.setCurrentIndex(index)
            self.CementtableView.edit(index)
            self.CementMode.submit()
        except BaseException:
            QMessageBox.about(QWidget(), '失败', '失败')

    def addMix(self):
        try:
            row = self.MixMode.rowCount()
            self.MixMode.insertRow(row)
            index = self.MixMode.index(row, 0)
            self.MixtableView.setCurrentIndex(index)
            self.MixtableView.edit(index)
            self.MixMode.submit()
        except BaseException:
            QMessageBox.about(QWidget(), '失败', '失败')

    def CementInfo(self):
        from DateBase.insert_value import insert_cement_attribute_data
        try:
            filePath, filetype = QFileDialog.getOpenFileName(
                QWidget(), "选取水泥资料一览表", "../", "Text Files (*.xlsx)")
            if filePath == '':
                raise Exception('请选择文件夹')
            self.CementInfoPath.setText(filePath)
            filename = self.CementInfoPath.text()
            print(filename)
            insert_cement_attribute_data(filename)
            QMessageBox.information(QWidget(), "成功", "水泥资料一览表导入成功")
        except BaseException:
            QMessageBox.information(
                QWidget(), "错误", "添加失败！！！！！\n请检查水泥资料一览表表格的数据格式。\n添加失败！！！！")

    def ConMixInsert(self):
        from DateBase.insert_value import insert_concrete_mix
        try:
            filePath, filetype = QFileDialog.getOpenFileName(
                QWidget(), "选取配合比选用汇总表", "../", "Text Files (*.xlsx)")
            if filePath == '':
                raise Exception('请选择文件夹')
            self.MixPoportionPath.setText(filePath)
            filename = self.MixPoportionPath.text()
            insert_concrete_mix(filename)
            QMessageBox.information(QWidget(), "成功", "配合比选用汇总表导入成功")
        except BaseException:
            QMessageBox.information(
                QWidget(), "错误", "添加失败！！！！！\n请检查配合比选用汇总表格的数据格式。\n添加失败！！！！")

    def ChoicConcreteUsageRecord(self):
        try:
            filePath, filetype = QFileDialog.getOpenFileName(
                QWidget(), "选取水泥使用一览表", "../", "Text Files (*.xlsx)")
            if filePath == '':
                raise Exception('请选择文件夹')
            self.ConcreteUsageRecordPath.setText(filePath)
            self.OutPutButton.setEnabled(True)
            self.OutPutVoidButton.setEnabled(True)
        except BaseException:
            pass

    def ChoicSQLPath(self):
        try:
            filePath = QFileDialog.getExistingDirectory(
                QWidget(), "选取生成表格的文件夹", "../", )
            print(filePath)
            if filePath == '':
                print('没有选中文件')
                raise Exception('请选择文件夹')
            self.OutSQLButton_3.setEnabled(True)
            self.OutSQLPath.setText(filePath)
        except BaseException:
            pass

    def OutPutSQL(self):
        try:
            art = session.query(CementAttributeDatum).all()
            mix = session.query(ConcreteMix).all()
            wb = Workbook()
            ws = wb.active
            ws['A1'] = '水泥资料一览表'
            ws.merge_cells('A1:N1')
            ws['A2'] = '进场日期'
            ws['B2'] = '水泥品种'
            ws['C2'] = '生产厂家'
            ws['D2'] = '生产日期'
            ws['E2'] = '编号'
            ws['F2'] = '数量(T)'
            ws['G2'] = '安定性'
            ws['H2'] = '初凝'
            ws['I2'] = '终凝'
            ws['J2'] = 'R3抗压'
            ws['K2'] = 'R28抗压'
            ws['L2'] = 'R3抗折'
            ws['M2'] = 'R28抗折'
            ws['N2'] = '是否优先'
            for i in range(len(art)):
                one_art = art[i]
                ws['A' + str(i + 3)] = one_art.ArrivalTime
                ws['B' + str(i + 3)] = one_art.CementVariety
                ws['C' + str(i + 3)] = one_art.Manufacturer
                ws['D' + str(i + 3)] = one_art.ProductionDate
                ws['E' + str(i + 3)] = one_art.CementId
                ws['F' + str(i + 3)] = one_art.CementNumber
                if one_art.IsStability == 1:
                    ws['G' + str(i + 3)] = '合格'
                else:
                    ws['G' + str(i + 3)] = '不合格'
                ws['H' + str(i + 3)] = one_art.InitialTime
                ws['I' + str(i + 3)] = one_art.FinalTime
                ws['J' + str(i + 3)] = one_art.R3_Compression
                ws['K' + str(i + 3)] = one_art.R28_Compression
                ws['L' + str(i + 3)] = one_art.R3_Bending
                ws['M' + str(i + 3)] = one_art.R28_Bending
                ws['N' + str(i + 3)] = one_art.PriorityLevel
            wb.save(self.OutSQLPath.text() + '/水泥一览表test.xlsx')
            # SQL配合比表
            wb = Workbook()
            ws = wb.active
            ws['A1'] = '佛山市三水区建友混凝土有限公司'
            ws['A2'] = '配合比选用汇总表'
            ws.merge_cells('A1:W1')
            ws.merge_cells('A2:W2')
            ws['A4'] = '名称'
            ws['B4'] = '配合比编号'
            ws['C4'] = '强度等级 '
            ws.merge_cells('C4:D4')
            ws['E4'] = '抗渗等级'
            ws['F4'] = '膨胀'
            ws['G4'] = '配合比编号'
            ws['H4'] = '坍落度'
            ws['I4'] = '标准差(MPa)'
            ws['J4'] = '配制强度(MPa)W'
            ws['K4'] = '水W'
            ws['L4'] = '水泥C'
            ws['M4'] = '粉煤灰F'
            ws['N4'] = '砂S'
            ws['O4'] = '石G'
            ws['P4'] = '水胶比A/P'
            ws['Q4'] = '砂率 BS'
            ws['R4'] = '外加剂掺量A%'
            ws['S4'] = '外加剂用量LS-JS(B)'
            ws['T4'] = '膨胀剂用量'
            ws['U4'] = '质量密度  (容重）Mcp'
            ws['V4'] = '初凝时间'
            ws['W4'] = '终凝时间'
            for i in range(len(mix)):
                ws['C' + str(i + 5)] = 'C'
                ws['A' + str(i + 5)] = mix[i].ConcreteName
                ws['B' + str(i + 5)] = mix[i].MixRatioID
                ws['D' + str(i + 5)] = mix[i].StrengthLevel
                ws['E' + str(i + 5)] = mix[i].ImperLevel
                if mix[i].ImperLevel is None:
                    ws['E' + str(i + 5)] = '/'
                ws['F' + str(i + 5)] = mix[i].SwellLevel
                if mix[i].SwellLevel is None:
                    ws['F' + str(i + 5)] = '/'
                ws['G' + str(i + 5)] = mix[i].MixRatioName
                ws['H' + str(i + 5)] = mix[i].SlumpNum
                ws['I' + str(i + 5)] = mix[i].StandardDeviation
                ws['J' + str(i + 5)] = mix[i].ConcreteStrengh
                ws['K' + str(i + 5)] = mix[i].WaterNum
                ws['L' + str(i + 5)] = mix[i].CementNum
                ws['N' + str(i + 5)] = mix[i].FlyashNum
                ws['M' + str(i + 5)] = mix[i].SandNum
                ws['O' + str(i + 5)] = mix[i].GravelNum
                ws['P' + str(i + 5)] = mix[i].CementRatio
                ws['Q' + str(i + 5)] = mix[i].SandRatio
                ws['R' + str(i + 5)] = mix[i].AdmixtureAmount
                ws['S' + str(i + 5)] = mix[i].AdmixtureNum
                ws['T' + str(i + 5)] = mix[i].SwellingNum
                ws['U' + str(i + 5)] = mix[i].MassDensity
                ws['V' + str(i + 5)] = mix[i].InitialTime
                ws['W' + str(i + 5)] = mix[i].FinalTime
            wb.save(self.OutSQLPath.text() + '/配合比表test.xlsx')
            QMessageBox.information(QWidget(), "成功", "请在刚刚选中的目录下查看文件")
        except BaseException:
            QMessageBox.information(
                QWidget(), "错误", "添加失败！！！！！\n请检查配合比选用汇总表格的数据格式。\n添加失败！！！！")

    def ouputexcel(self):
        try:
            books = []
            book = load_workbook(self.ConcreteUsageRecordPath.text())
            print('成功读取水泥使用记录表')
            # 生成12份表格
            book1, book7 = ConUseBuyRecord(book)
            print("17")
            book2, book3, book4 = ConStrengReport(book)
            print("234")
            book5 = CreateConFacyoryCard(book, book1)
            print("5")
            book6 = CreateSlumpCard(book5)
            print("6")
            book8 = CreteConUseProve(book)
            print("8")
            book9 = ConQualityRecord(book, book3, book7)
            print("9")
            book10 = PermeabilityTestReport(book, book7)
            print("10")
            books.append(book1)
            books.append(book2)
            books.append(book3)
            books.append(book4)
            books.append(book5)
            books.append(book6)
            books.append(book7)
            books.append(book8)
            books.append(book9)
            books.append(book10)
        except BaseException:
            print("读取失败")
        try:
            # 文件名处理
            fileNames = [
                '1、水泥购进、使用一览表',
                '2、混凝土试件抗压强度检验报告',
                '3、混凝土试块强度试验结果汇总表',
                '4、标准养护混凝土抗压强度计算表',
                '5、建友商品混凝土出厂合格证',
                '6、混凝土坍落度验收表',
                '7、混凝土配合比设计报告',
                '8、工地使用预拌混凝土证明书',
                '9、混凝土搅拌质量记录表',
                '10、抗渗性能检测报告',
                '11、施工配合比']
            filePath = QFileDialog.getExistingDirectory(
                QWidget(), "选取生成表格的文件夹", "C:/", )
            print(filePath)
            if filePath == '':
                print('没有选中文件')
                raise Exception('请选择文件夹')
            filePaths = []  # 存放文件路劲
            for i in fileNames:
                filePaths.append(filePath + '/' + i + '.xlsx')
            for j in range(1, 11):
                books[j - 1].save(filePaths[j - 1])
            # 记录
            filePaths = []
            filePath = his()
            for i in fileNames:
                filePaths.append(filePath + '/' + i + '.xlsx')
            print(filePaths)
            for j in range(1, 11):
                books[j - 1].save(filePaths[j - 1])
            QMessageBox.information(QWidget(), "成功导出", "请在您刚刚选定的文件夹内查看生成的文件")
        except BaseException:
            QMessageBox.information(QWidget(), "错误", "添加失败！！！！！\n添加失败！！！！")

    def outputvoidexcel(self):
        from Excel.CreateMode.CreateConFactoryCardMode17 import ConUseBuyRecordMode
        from Excel.CreateMode.CreateConStrentReportMode234 import ConStrengReportMode
        from Excel.CreateMode.CreateConFactoryCardMode56 import CreateConFacyoryCardMode, CreateSlumpCardMode
        from Excel.CreateMode.ConUseProveMode8 import CreteConUseProveMode
        from Excel.CreateMode.ConQualityRecordMode9 import ConQualityRecordMode
        from Excel.CreateMode.PermeabilityTestReportMode10 import PermeabilityTestReportMode
        try:
            books = []
            book = load_workbook(self.ConcreteUsageRecordPath.text())
            print('成功读取水泥使用记录表')
            # 生成12份表格
            book1, book7 = ConUseBuyRecordMode(book)
            print("17")
            book2, book3, book4 = ConStrengReportMode(book)
            print("234")
            book5 = CreateConFacyoryCardMode(book)
            print("5")
            book6 = CreateSlumpCardMode(book5)
            print("6")
            book8 = CreteConUseProveMode(book)
            print("8")
            book9 = ConQualityRecordMode(book)
            print("9")
            book10 = PermeabilityTestReportMode(book)
            print("10")
            books.append(book1)
            books.append(book2)
            books.append(book3)
            books.append(book4)
            books.append(book5)
            books.append(book6)
            books.append(book7)
            books.append(book8)
            books.append(book9)
            books.append(book10)
        except BaseException:
            print("读取失败")
        try:
            # 文件名处理
            fileNames = [
                '1、水泥购进、使用一览表',
                '2、混凝土试件抗压强度检验报告',
                '3、混凝土试块强度试验结果汇总表',
                '4、标准养护混凝土抗压强度计算表',
                '5、建友商品混凝土出厂合格证',
                '6、混凝土坍落度验收表',
                '7、混凝土配合比设计报告',
                '8、工地使用预拌混凝土证明书',
                '9、混凝土搅拌质量记录表',
                '10、抗渗性能检测报告',
                '11、施工配合比']
            filePath = QFileDialog.getExistingDirectory(
                QWidget(), "选取生成表格的文件夹", "C:/", )
            print(filePath)
            if filePath == '':
                print('没有选中文件')
                raise Exception('请选择文件夹')
            filePaths = []  # 存放文件路劲
            for i in fileNames:
                filePaths.append(filePath + '/' + i + '空白.xlsx')
            for j in range(1, 11):
                books[j - 1].save(filePaths[j - 1])
            QMessageBox.information(QWidget(), "成功导出", "请在您刚刚选定的文件夹内查看生成的文件")
        except BaseException:
            QMessageBox.information(QWidget(), "错误", "添加失败！！！！！\n添加失败！！！！")


if __name__ == "__main__":
    pass
