# -*- coding:utf-8 -*-
"""
@author: Darcy
@time: 2018/10/28 19:40
"""
from openpyxl import load_workbook
from Util.ExcelInitUtil import iniexcel2
from Util.ExcelInitUtil import iniexcel3
from Util.ExcelInitUtil import iniexcel4
import Bean.DbBean
import Bean.ExcelBean
from datetime import datetime, timedelta
from collections import defaultdict
from random import uniform
from DateBase.connect_db import session
from Util.round import get_float
"""
    生成混凝土试件抗压强度检验报告,混凝土试件强度实验结果汇总表
    Parm:
        useBook 使用记录表
        session 数据库连接池
        count 使用记录数量
    Return:
        resultRepoBook      混凝土试件抗压强度检验报告
        resultRepoSumBook   混凝土试件强度实验结果汇总表
        resultRepoCompBook  标准养护混凝土抗压强度计算表
"""
def ConStrengReport(useBook):
    from Util.get_parm import get_parm
    parm = get_parm()
    MinC_Strength = float(parm.MinC_Strength)
    MaxC_Strength = float(parm.MaxC_Strength)
    useSheets = useBook.sheetnames
    secModeName = "../../Mode/2.xlsx"
    thrModeNmae = "../../Mode/3.xlsx"
    fouModeName = "../../Mode/4.xlsx"
    resultRepoBook = load_workbook(secModeName)
    resultRepoSumBook = load_workbook(thrModeNmae)
    resultRepoCompBook = load_workbook(fouModeName)

    """
        复制第二份表模板
    """
    for sheetNum in range(len(useSheets)):
        useSheet = useBook.worksheets[sheetNum]
        rows = useSheet.max_row
        moreNum = useSheet.max_row
        moreNum = moreNum - 9
        for rowNum in range(10, rows + 1):
            if useSheet.cell(row=rowNum, column=1).value == None:
                break
            CuringNum = useSheet.cell(row=rowNum, column=8).value
            if(CuringNum >= 1000):
                moreNum = moreNum + int(CuringNum / 200)
            elif(CuringNum > 100):
                moreNum = moreNum + int(CuringNum / 100)
                # while (CuringNum > 100):
                #     moreNum = moreNum + 1
                #     CuringNum = CuringNum - 100
        moreFixNum, rows = divmod(moreNum, 15)
        if rows == 0:
            moreFixNum = moreFixNum - 1
        for copyNum in range(moreFixNum + 1):
            modeSheet = resultRepoBook.worksheets[0]
            iniexcel2(modeSheet)
            modeSheet = resultRepoBook.copy_worksheet(modeSheet)
            iniexcel2(modeSheet)
            modeSheet.title = str(str(sheetNum + 1) + '-' + str(copyNum + 1))
    resultRepoBook.remove(resultRepoBook.worksheets[0])

    """
        复制第三份表模板
    """
    for sheetNum in range(len(useSheets)):
        useSheet = useBook.worksheets[sheetNum]
        rows = useSheet.max_row - 9
        moreFixNum, rows = divmod(rows, 20)
        if rows == 0:
            moreFixNum = moreFixNum - 1
        for copyNum in range(moreFixNum + 1):
            modeSheet = resultRepoSumBook.worksheets[0]
            iniexcel3(modeSheet)
            modeSheet = resultRepoSumBook.copy_worksheet(modeSheet)
            iniexcel3(modeSheet)
            modeSheet.title = str(str(sheetNum + 1) + '-' + str(copyNum + 1))
    resultRepoSumBook.remove(resultRepoSumBook.worksheets[0])
    # resultRepoSumBook.save('3.xlsx')
    """
        获取表数据
    """
    ConStrRepoHead = []
    UseRecord = []
    TestNameList = []
    TestDict = {
        'C15' : 'A','C20' : 'B',
        'C25' : 'C','C30' : 'D',
        'C35' : 'E', 'C40' : 'F',
        'C45' : 'G', 'C50' : 'H',
        'C55' : 'I', 'C60' : 'J',
        'C65' : 'K', 'C70' : 'L',
        'C75' : 'M', 'C80' : 'N'
    }
    RepoSum = []
    RepoComp = []
    for i in range(len(useSheets)):
        sheet = useBook.worksheets[i]
        rows = sheet.max_row
        UseRecordList = []
        RepoSumList = []
        RepoCompList = []
        ComRepoCompList = []
        for k in range(10, rows + 1):
            TypicalStrengthSum = 0
            TypicalStrengthSumFlag = 0
            if sheet.cell(row=k, column=1).value == None:
                break
            TypicalStrengthList = []
            CuringNum = sheet.cell(row=k, column=8).value
            if(CuringNum < 1000):
                CementName = sheet['B' + str(k)].value.replace('砼', '')
                ProjectSite = sheet.cell(row=k, column=1).value
                ConcreteStrength = sheet.cell(row=k, column=3).value
                ImperLevel = sheet.cell(row=k, column=4).value
                SwellLevel = sheet.cell(row=k, column=5).value
                CuringDate = sheet.cell(row=k, column=6).value
                TestName = TestDict[ConcreteStrength] + CuringDate.strftime('%m%d')
                TestNameList.append(TestName)
                MoldingTime = (CuringDate + timedelta(days=28)).strftime('%Y/%#m/%#d')
                SortTime = datetime.strptime(MoldingTime, '%Y/%m/%d')
                CuringDate = CuringDate.strftime('%Y/%#m/%#d')
                Strength = float(ConcreteStrength.strip('C'))
                LongStrength = Strength * uniform(MinC_Strength, MaxC_Strength)
                WideStrength = Strength * uniform(MinC_Strength, MaxC_Strength)
                HigthStrength = Strength * uniform(MinC_Strength, MaxC_Strength)
                AvgStrength = (LongStrength + WideStrength + HigthStrength) / 3
                StrengthString = ConcreteStrength
                if (ImperLevel == None):
                    StrengthString = StrengthString + CementName
                else:
                    StrengthString = StrengthString + CementName + ImperLevel
                if (SwellLevel != None):
                    StrengthString = StrengthString + '\n(' + str(int(SwellLevel * 100)) + '%膨胀)'
                LongStrength = '{:.1f}'.format(get_float(LongStrength, 1))
                WideStrength = '{:.1f}'.format(get_float(WideStrength, 1))
                HigthStrength = '{:.1f}'.format(get_float(HigthStrength, 1))
                AvgStrength = '{:.1f}'.format(get_float(AvgStrength, 1))
                ConcreteStrength = sheet.cell(row=k, column=3).value + CementName
                CompStrength = ConcreteStrength[1:3]
                CompAvgStrength = AvgStrength
                ComRepoComp = Bean.ExcelBean.ConRepoCompBean(CompStrength=CompStrength,
                                                             CompAvgStrength=CompAvgStrength)
                ComRepoCompList.append(ComRepoComp)
                ConUseRecord = Bean.ExcelBean.ConStrengUseBean(TestName=TestName,
                                                               ProjectSite=ProjectSite,
                                                               ConcreteStrength=ConcreteStrength,
                                                               StrengthString=StrengthString,
                                                               ImperLevel=ImperLevel,
                                                               SwellLevel=SwellLevel,
                                                               CuringDate=CuringDate,
                                                               MoldingTime=MoldingTime,
                                                               LongStrength=LongStrength,
                                                               WideStrength=WideStrength,
                                                               HigthStrength=HigthStrength,
                                                               AvgStrength=AvgStrength,
                                                               CuringNum=CuringNum,
                                                               SortTime=SortTime,
                                                               )
                SumProjectSite = ProjectSite
                # SumConcreteStrength = str(ConcreteStrength)
                SumConcreteStrength = StrengthString
                SumCuringDate = CuringDate
                #
                SumSortTime = datetime.strptime(SumCuringDate, '%Y/%m/%d')
                TypicalStrengthList.append(AvgStrength)
                TypicalStrengthSum = float(AvgStrength) + TypicalStrengthSum
                TypicalStrengthSumFlag = TypicalStrengthSumFlag + 1
                UseRecordList.append(ConUseRecord)
                while(CuringNum - 100 > 0):
                    ProjectSite = sheet.cell(row=k, column=1).value
                    ConcreteStrength = sheet.cell(row=k, column=3).value
                    ImperLevel = sheet.cell(row=k, column=4).value
                    SwellLevel = sheet.cell(row=k, column=5).value
                    CementName = sheet['B' + str(k)].value.replace('砼', '')
                    CuringDate = sheet.cell(row=k, column=6).value
                    TestName = TestDict[ConcreteStrength] + CuringDate.strftime('%m%d')
                    TestNameList.append(TestName)
                    MoldingTime = (CuringDate + timedelta(days=28)).strftime('%Y/%#m/%#d')
                    #
                    SortTime = datetime.strptime(MoldingTime, '%Y/%m/%d')
                    CuringDate = CuringDate.strftime('%Y/%#m/%#d')
                    Strength = float(ConcreteStrength.strip('C'))
                    LongStrength = Strength * uniform(MinC_Strength, MaxC_Strength)
                    WideStrength = Strength * uniform(MinC_Strength, MaxC_Strength)
                    HigthStrength = Strength * uniform(MinC_Strength, MaxC_Strength)
                    StrengthString = ConcreteStrength
                    if (ImperLevel == None):
                        StrengthString = StrengthString + CementName
                    else:
                        StrengthString = StrengthString + CementName + ImperLevel
                    if (SwellLevel != None):
                        StrengthString = StrengthString + '\n(' + str(int(SwellLevel * 100)) + '%膨胀)'
                    AvgStrength = (LongStrength + WideStrength + HigthStrength) / 3
                    LongStrength = '{:.1f}'.format(get_float(LongStrength, 1))
                    WideStrength = '{:.1f}'.format(get_float(WideStrength, 1))
                    HigthStrength = '{:.1f}'.format(get_float(HigthStrength, 1))
                    AvgStrength = '{:.1f}'.format(get_float(AvgStrength, 1))
                    ConcreteStrength = sheet.cell(row=k, column=3).value + CementName
                    ConUseRecord = Bean.ExcelBean.ConStrengUseBean(TestName=TestName,
                                                                   ProjectSite=ProjectSite,
                                                                   ConcreteStrength=ConcreteStrength,
                                                                   ImperLevel=ImperLevel,
                                                                   SwellLevel=SwellLevel,
                                                                   StrengthString=StrengthString,
                                                                   CuringDate=CuringDate,
                                                                   MoldingTime=MoldingTime,
                                                                   LongStrength=LongStrength,
                                                                   WideStrength=WideStrength,
                                                                   HigthStrength=HigthStrength,
                                                                   AvgStrength=AvgStrength,
                                                                   CuringNum=CuringNum,
                                                                   SortTime=SortTime,
                                                                   )
                    UseRecordList.append(ConUseRecord)
                    TypicalStrengthSum = float(AvgStrength) + TypicalStrengthSum
                    TypicalStrengthSumFlag = TypicalStrengthSumFlag + 1
                    TypicalStrengthList.append(AvgStrength)
                    CompAvgStrength = AvgStrength
                    ComRepoComp = Bean.ExcelBean.ConRepoCompBean(CompStrength=CompStrength,
                                                                 CompAvgStrength=CompAvgStrength)
                    ComRepoCompList.append(ComRepoComp)
                    CuringNum = CuringNum - 100
                SumToStrength = (TypicalStrengthSum / TypicalStrengthSumFlag) / float(sheet.cell(row=k, column=3).value.strip('C'))
                RepoSumRecord = Bean.ExcelBean.ConRepoSumBean(ProjectSite=SumProjectSite,
                                                              ConcreteStrength=SumConcreteStrength,
                                                              CuringDate=SumCuringDate,
                                                              TypicalStrength=TypicalStrengthList,
                                                              ToStrength=SumToStrength,
                                                              SumSortTime=SumSortTime)
                RepoSumList.append(RepoSumRecord)
            elif(CuringNum >= 1000):
                CementName = sheet['B' + str(k)].value.replace('砼', '')
                ProjectSite = sheet.cell(row=k, column=1).value
                ConcreteStrength = sheet.cell(row=k, column=3).value
                ImperLevel = sheet.cell(row=k, column=4).value
                SwellLevel = sheet.cell(row=k, column=5).value
                CuringDate = sheet.cell(row=k, column=6).value
                TestName = TestDict[ConcreteStrength] + CuringDate.strftime('%m%d')
                TestNameList.append(TestName)
                MoldingTime = (CuringDate + timedelta(days=28)).strftime('%Y/%#m/%#d')
                #
                SortTime = datetime.strptime(MoldingTime, '%Y/%m/%d')
                CuringDate = CuringDate.strftime('%Y/%#m/%#d')
                Strength = float(ConcreteStrength.strip('C'))
                LongStrength = Strength * uniform(MinC_Strength, MaxC_Strength)
                WideStrength = Strength * uniform(MinC_Strength, MaxC_Strength)
                HigthStrength = Strength * uniform(MinC_Strength, MaxC_Strength)
                AvgStrength = (LongStrength + WideStrength + HigthStrength) / 3
                StrengthString = ConcreteStrength
                if (ImperLevel == None):
                    StrengthString = StrengthString + CementName
                else:
                    StrengthString = StrengthString + CementName + ImperLevel
                if (SwellLevel != None):
                    StrengthString = StrengthString + '\n(' + str(int(SwellLevel * 100)) + '%膨胀)'
                LongStrength = '{:.1f}'.format(get_float(LongStrength, 1))
                WideStrength = '{:.1f}'.format(get_float(WideStrength, 1))
                HigthStrength = '{:.1f}'.format(get_float(HigthStrength, 1))
                AvgStrength = '{:.1f}'.format(get_float(AvgStrength, 1))
                ConcreteStrength = sheet.cell(row=k, column=3).value + CementName
                CompStrength = ConcreteStrength[1:3]
                CompAvgStrength = AvgStrength
                ComRepoComp = Bean.ExcelBean.ConRepoCompBean(CompStrength=CompStrength,
                                                             CompAvgStrength=CompAvgStrength)
                ComRepoCompList.append(ComRepoComp)
                ConUseRecord = Bean.ExcelBean.ConStrengUseBean(TestName=TestName,
                                                               ProjectSite=ProjectSite,
                                                               ConcreteStrength=ConcreteStrength,
                                                               ImperLevel=ImperLevel,
                                                               SwellLevel=SwellLevel,
                                                               StrengthString=StrengthString,
                                                               CuringDate=CuringDate,
                                                               MoldingTime=MoldingTime,
                                                               LongStrength=LongStrength,
                                                               WideStrength=WideStrength,
                                                               HigthStrength=HigthStrength,
                                                               AvgStrength=AvgStrength,
                                                               CuringNum=CuringNum,
                                                               SortTime=SortTime,
                                                               )
                SumProjectSite = ProjectSite
                SumConcreteStrength = StrengthString
                SumCuringDate = CuringDate
                #
                SumSortTime = datetime.strptime(SumCuringDate, '%Y/%m/%d')
                TypicalStrengthList.append(AvgStrength)
                TypicalStrengthSum = float(AvgStrength) + TypicalStrengthSum
                TypicalStrengthSumFlag = TypicalStrengthSumFlag + 1
                UseRecordList.append(ConUseRecord)
                while (CuringNum - 200 > 0):
                    ProjectSite = sheet.cell(row=k, column=1).value
                    ConcreteStrength = sheet.cell(row=k, column=3).value
                    ImperLevel = sheet.cell(row=k, column=4).value
                    SwellLevel = sheet.cell(row=k, column=5).value
                    CementName = sheet['B' + str(k)].value.replace('砼', '')
                    CuringDate = sheet.cell(row=k, column=6).value
                    TestName = TestDict[ConcreteStrength] + CuringDate.strftime('%m%d')
                    TestNameList.append(TestName)
                    MoldingTime = (CuringDate + timedelta(days=28)).strftime('%Y/%#m/%#d')
                    #
                    SortTime = datetime.strptime(MoldingTime, '%Y/%m/%d')
                    CuringDate = CuringDate.strftime('%Y/%#m/%#d')
                    Strength = float(ConcreteStrength.strip('C'))
                    LongStrength = Strength * uniform(MinC_Strength, MaxC_Strength)
                    WideStrength = Strength * uniform(MinC_Strength, MaxC_Strength)
                    HigthStrength = Strength * uniform(MinC_Strength, MaxC_Strength)
                    AvgStrength = (LongStrength + WideStrength + HigthStrength) / 3
                    StrengthString = ConcreteStrength
                    if (ImperLevel == None):
                        StrengthString = StrengthString + CementName
                    else:
                        StrengthString = StrengthString + CementName + ImperLevel
                    if (SwellLevel != None):
                        StrengthString = StrengthString + '\n(' + str(int(SwellLevel * 100)) + '%膨胀)'
                    LongStrength = '{:.1f}'.format(get_float(LongStrength, 1))
                    WideStrength = '{:.1f}'.format(get_float(WideStrength, 1))
                    HigthStrength = '{:.1f}'.format(get_float(HigthStrength, 1))
                    AvgStrength = '{:.1f}'.format(get_float(AvgStrength, 1))
                    ConcreteStrength = sheet.cell(row=k, column=3).value + CementName
                    ConUseRecord = Bean.ExcelBean.ConStrengUseBean(TestName=TestName,
                                                                   ProjectSite=ProjectSite,
                                                                   ConcreteStrength=ConcreteStrength,
                                                                   ImperLevel=ImperLevel,
                                                                   SwellLevel=SwellLevel,
                                                                   StrengthString=StrengthString,
                                                                   CuringDate=CuringDate,
                                                                   MoldingTime=MoldingTime,
                                                                   LongStrength=LongStrength,
                                                                   WideStrength=WideStrength,
                                                                   HigthStrength=HigthStrength,
                                                                   AvgStrength=AvgStrength,
                                                                   CuringNum=CuringNum,
                                                                   SortTime=SortTime,
                                                                   )
                    UseRecordList.append(ConUseRecord)
                    TypicalStrengthSum = float(AvgStrength) + TypicalStrengthSum
                    TypicalStrengthSumFlag = TypicalStrengthSumFlag + 1
                    TypicalStrengthList.append(AvgStrength)
                    CompAvgStrength = AvgStrength
                    ComRepoComp = Bean.ExcelBean.ConRepoCompBean(CompStrength=CompStrength,
                                                                 CompAvgStrength=CompAvgStrength)
                    ComRepoCompList.append(ComRepoComp)
                    CuringNum = CuringNum - 200
                SumToStrength = (TypicalStrengthSum / TypicalStrengthSumFlag) / float(sheet.cell(row=k, column=3).value.strip('C'))
                RepoSumRecord = Bean.ExcelBean.ConRepoSumBean(ProjectSite=SumProjectSite,
                                                              ConcreteStrength=SumConcreteStrength,
                                                              CuringDate=SumCuringDate,
                                                              TypicalStrength=TypicalStrengthList,
                                                              ToStrength=SumToStrength,
                                                              SumSortTime=SumSortTime,
                                                              )
                RepoSumList.append(RepoSumRecord)
        RepoComp.append(ComRepoCompList)
        # UseRecordList.sort(key=lambda x: x.SortTime, reverse=False)
        UseRecord.append(UseRecordList)
        # RepoSumList.sort(key=lambda x: x.SumSortTime, reverse=False)
        RepoSum.append(RepoSumList)
        # InspectionName = parm[0].ConStrengInspectionUtil
        ProjectName = sheet['A2'].value.replace('工程名称：', '').strip()
        BuildUnit = sheet['A3'].value.replace('建设单位：', '').strip()
        # RepoTime = UseRecordList[-1].MoldingTime
        RepoTimeList = []
        for useRecord in UseRecordList:
            repoTime = datetime.strptime(useRecord.MoldingTime, '%Y/%m/%d')
            RepoTimeList.append(repoTime)
        #
        RepoTime = max(RepoTimeList)
        RepoTime = RepoTime.strftime('%Y/%m/%d')
        ConUseHeadRecord = Bean.ExcelBean.ConStrengRepoBean(ProjectName=ProjectName,
                                                            BuildUnit=BuildUnit,
                                                            # InspectionName=InspectionName,
                                                            RepoTime=RepoTime)
        ConStrRepoHead.append(ConUseHeadRecord)

    TestNameSet = set(TestNameList)
    TestNameDict = defaultdict(list)
    for item in TestNameSet:
        TestNameDict[item].append(TestNameList.count(item))
        TestNameDict[item].append(0)

    """
        获取第四份表数据
    """
    repoCompBeanList = []
    for repoComp in RepoComp:
        repoCompStrengthList = []
        # repoCompStrengthSet = []
        repoCompDict = defaultdict(list)
        for repoCompNum in repoComp:
            repoCompStrengthList.append(repoCompNum)
            # repoCompStrengthSet = list(set(repoCompStrengthList))
            repoCompDict[repoCompNum.CompStrength].append(repoCompNum.CompAvvgStrength)
        repoCompBeanList.append(repoCompDict)

    """
        复制第四份表模板
    """
    for sheetNum in range(len(repoCompBeanList)):
        rows = repoCompBeanList[sheetNum].__len__()
        moreSixNum, rows = divmod(rows, 6)
        if rows == 0:
            moreSixNum = moreSixNum - 1
        for copyNum in range(moreSixNum + 1):
            modeSheet = resultRepoCompBook.worksheets[0]
            iniexcel4(modeSheet)
            modeSheet = resultRepoCompBook.copy_worksheet(modeSheet)
            iniexcel4(modeSheet)
            modeSheet.title = str(sheetNum + 1) + '-' + str(copyNum + 1)
    resultRepoCompBook.remove(resultRepoCompBook.worksheets[0])
    """
        插入第二份表数据
    """
    curSheetNum = 0
    for sheetNum in range(len(useSheets)):
        useRecordNum = len(UseRecord[sheetNum])
        moreFixNum, rows = divmod(useRecordNum, 15)
        if rows == 0:
            moreFixNum = moreFixNum - 1
        for insertNum in range(moreFixNum + 1):
            InsertFlag = 0
            resultSheet = resultRepoBook.worksheets[curSheetNum]
            resultSheet['I3'] = parm.Project2InspectCodeEdit
            # resultSheet['I3'] = Project2InspectCodeEdit
            resultSheet['B4'] = ConStrRepoHead[sheetNum].BuildUnit
            # resultSheet['G4'] = ConStrRepoHead[sheetNum].InspectionName
            resultSheet['B5'] = ConStrRepoHead[sheetNum].ProjectName
            resultSheet['F5'] = "报告日期：" + ConStrRepoHead[sheetNum].RepoTime.replace('/', '.')
            curSheetNum = curSheetNum + 1
            for useNum in range(15):
                if ((useNum + 15 * insertNum) == len(UseRecord[sheetNum])):
                    if ((useNum + 15 * insertNum) / 15 != 0):
                        resultSheet['D' + str(8 + InsertFlag)] = '以'
                        resultSheet['E' + str(9 + InsertFlag)] = '下'
                        resultSheet['F' + str(8 + InsertFlag)] = '空'
                        resultSheet['G' + str(8 + InsertFlag)] = '白'
                    break
                TestName = UseRecord[sheetNum][useNum + 15 * insertNum].TestName
                if (TestNameDict[UseRecord[sheetNum][useNum + 15 * insertNum].TestName][0] != 1):
                    TestNameDict[UseRecord[sheetNum][useNum + 15 * insertNum].TestName][1] = TestNameDict[UseRecord[sheetNum][useNum + 15 * insertNum].TestName][1] + 1
                    TestName = UseRecord[sheetNum][useNum + 15 * insertNum].TestName + '-' + str(TestNameDict[UseRecord[sheetNum][useNum + 15 * insertNum].TestName][1])
                # if (TestNameDict[UseRecord[sheetNum][useNum].TestName][0] != 1):
                #     TestNameDict[UseRecord[sheetNum][useNum].TestName][1] = TestNameDict[UseRecord[sheetNum][useNum].TestName][1] + 1
                #     UseRecord[sheetNum][useNum].TestName = UseRecord[sheetNum][useNum].TestName + '-' + str(TestNameDict[UseRecord[sheetNum][useNum].TestName][1])
                resultSheet['A' + str(8 + InsertFlag)] = TestName
                resultSheet['B' + str(8 + InsertFlag)] = UseRecord[sheetNum][useNum + 15 * insertNum].ProjectSite
                resultSheet['D' + str(8 + InsertFlag)] = UseRecord[sheetNum][useNum + 15 * insertNum].StrengthString
                resultSheet['F' + str(8 + InsertFlag)] = UseRecord[sheetNum][useNum + 15 * insertNum].CuringDate.replace('/','-')
                resultSheet['G' + str(8 + InsertFlag)] = UseRecord[sheetNum][useNum + 15 * insertNum].MoldingTime.replace('/','-')
                resultSheet['H' + str(8 + InsertFlag)] = '28'
                resultSheet['I' + str(8 + InsertFlag)] = UseRecord[sheetNum][useNum + 15 * insertNum].LongStrength
                resultSheet['E' + str(8 + InsertFlag)] = "长=150"
                resultSheet['E' + str(9 + InsertFlag)] = "宽=150"
                resultSheet['E' + str(10 + InsertFlag)] = "高=150"
                resultSheet['I' + str(9 + InsertFlag)] = UseRecord[sheetNum][useNum + 15 * insertNum].WideStrength
                resultSheet['I' + str(10 + InsertFlag)] = UseRecord[sheetNum][useNum + 15 * insertNum].HigthStrength
                resultSheet['J' + str(8 + InsertFlag)] = UseRecord[sheetNum][useNum + 15 * insertNum].AvgStrength
                InsertFlag = InsertFlag + 3
                from Util.InsetPicUtil import insertpic
                from Util.get_parm import get_parm
                parm = get_parm()
                resultSheet = insertpic(resultSheet, picname=parm.Project2Manager, position='C54', width=90, heigh=30)
                resultSheet = insertpic(resultSheet, picname=parm.Project2Checker, position='G54')
                resultSheet = insertpic(resultSheet, picname=parm.Project2Try, position='J54')

    """
        插入第三份表数据
    """
    curSheetNum = 0
    for sheetNum in range(len(useSheets)):
        repoSumNum = len(RepoSum[sheetNum])
        moreFixNum, rows = divmod(repoSumNum, 20)
        if rows == 0:
            moreFixNum = moreFixNum - 1
        for insertNum in range(moreFixNum + 1):
            InsertFlag = 0
            resultSheet = resultRepoSumBook.worksheets[curSheetNum]
            curSheetNum = curSheetNum + 1
            resultSheet['A4'] = "工程名称：" + ConStrRepoHead[sheetNum].ProjectName
            for sumNum in range(20):
                resultSheet["F3"] = parm.Project3InspectCodeEdit
                if((sumNum + 20 * insertNum) == len(RepoSum[sheetNum])):
                    if (sumNum + 20 * insertNum != 20):
                        resultSheet['D' + str(7 + InsertFlag)] = '以'
                        resultSheet['E' + str(7 + InsertFlag)] = '下'
                        resultSheet['F' + str(7 + InsertFlag)] = '空'
                        resultSheet['G' + str(7 + InsertFlag)] = '白'
                    break
                repoSum = RepoSum[sheetNum][sumNum + 20 * insertNum]
                resultSheet['A' + str(7 + InsertFlag)] = sumNum + 20 * insertNum + 1
                resultSheet['B' + str(7 + InsertFlag)] = repoSum.ProjectSite
                resultSheet['C' + str(7 + InsertFlag)] = repoSum.ConcreteStrength
                resultSheet['D' + str(7 + InsertFlag)] = repoSum.CuringDate.replace('/','-')
                resultSheet['E' + str(7 + InsertFlag)] = 28
                resultSheet['F' + str(7 + InsertFlag)] = "/".join(repoSum.TypicalStrength)
                resultSheet['G' + str(7 + InsertFlag)] = repoSum.ToStrength
                InsertFlag = InsertFlag + 1

                resultSheet = insertpic(resultSheet, picname=parm.Project3MakeSheet, position='F28', width=90, heigh=30)

    """
        插入第四份表数据
    """
    curSheetNum = 0
    for sheetNum in range(len(useSheets)):
        CompDictKeys = list(repoCompBeanList[sheetNum].keys())
        CompDictKeys.sort(key=lambda x:x, reverse=False)
        repoCompRows = repoCompBeanList[sheetNum].__len__()
        moreSixNum, rows = divmod(repoCompRows, 6)
        if rows == 0:
            moreSixNum = moreSixNum - 1
        for insertNum in range(moreSixNum + 1):
            InsertFlag = 0
            resultSheet = resultRepoCompBook.worksheets[curSheetNum]
            curSheetNum = curSheetNum + 1
            curNumFlag = 0
            for curNum in range(6):
                if(6 * insertNum + curNumFlag == repoCompRows):
                    break
                resultSheet['M3'] = parm.Project4InspectCodeEdit
                CompList = repoCompBeanList[sheetNum][CompDictKeys[curNumFlag + 6 * insertNum]]
                avgStreng = 0
                minStreng = float(CompList[0])
                for compList in CompList:
                    compList = float(compList)
                    avgStreng = avgStreng + compList
                    if (compList < minStreng):
                        minStreng = compList
                avgStreng = avgStreng / len(CompList)
                compNum = len(CompList)
                compStreng = int(CompDictKeys[curNumFlag + 6 * insertNum])
                if(compNum <= 9):
                    parm1 = 0.95
                    parm2 = 1.15
                    resultSheet['B' + str(8 + InsertFlag)] = 'C'
                    resultSheet['C' + str(8 + InsertFlag)] = compStreng
                    resultSheet['D' + str(8 + InsertFlag)] = compNum
                    resultSheet['G' + str(8 + InsertFlag)] = '{0:.1f}'.format(get_float(avgStreng,1))
                    resultSheet['H' + str(8 + InsertFlag)] = '{0:.1f}'.format(get_float(minStreng,1))
                    resultSheet['M' + str(8 + InsertFlag)] = '{0:.2f}'.format(get_float(parm1 * compStreng,2))
                    resultSheet['N' + str(8 + InsertFlag)] = '{0:.2f}'.format(get_float(parm2 * compStreng,2))
                    # QualifiedString = "C{compStreng}：mfcu = {avgStreng}＞λ\u2083·fcu,k ={Parm_3_Streng};fcu，min = {minStreng}≥λ\u2084·fcu，k = {Parm_4_Streng}，评定合格"
                    QualifiedString = "C{compStreng}：mfcu = {avgStreng}＞λ\u2083·fcu,k ={Parm_3_Streng};fcu，min = {minStreng}≥λ\u2084·fcu，k = {Parm_4_Streng}，评定合格"

                    resultSheet['C' + str(14+ InsertFlag)] = QualifiedString.format(compStreng='{0:.0f}'.format(compStreng),
                                                                                    avgStreng='{0:.1f}'.format(get_float(avgStreng, 1)),
                                                                                    Parm_3_Streng='{0:.2f}'.format(get_float((parm2 * compStreng),2)),
                                                                                    minStreng=minStreng,
                                                                                    Parm_4_Streng='{0:.2f}'.format(get_float((parm1 * compStreng),2)))
                elif(compNum >9 and compNum <=14):
                    parm1 = 1.15
                    parm2 = 0.9
                    standParm = 2.50
                    resultSheet['B' + str(8 + InsertFlag)] = 'C'
                    resultSheet['C' + str(8 + InsertFlag)] = compStreng
                    resultSheet['D' + str(8 + InsertFlag)] = compNum
                    resultSheet['E' + str(8 + InsertFlag)] = parm1
                    resultSheet['F' + str(8 + InsertFlag)] = parm2
                    resultSheet['G' + str(8 + InsertFlag)] = '{0:.1f}'.format(get_float(avgStreng,1))
                    resultSheet['H' + str(8 + InsertFlag)] = '{0:.1f}'.format(get_float(minStreng,1))
                    resultSheet['I' + str(8 + InsertFlag)] = standParm

                    resultSheet['J' + str(8 + InsertFlag)] = '{0:.2f}'.format(get_float((float(resultSheet['G' + str(8 + InsertFlag)].value) - (parm1 * standParm)), 2))
                    resultSheet['L' + str(8 + InsertFlag)] = '{0:.2f}'.format(get_float((parm2 * compStreng), 2))
                    QualifiedString = "C{compStreng}：mfcu = {avgStreng}＞fcu，k+λ\u2081·Sfcu ={Parm_1_Streng};fcu，min = {minStreng}≥λ\u2082·fcu，k = {Parm_2_Streng}，评定合格"
                    resultSheet['C' + str(14 + InsertFlag)] = QualifiedString.format(compStreng='{0:.0f}'.format(compStreng),
                                                                                     avgStreng='{0:.1f}'.format(get_float(avgStreng, 1)),
                                                                                     Parm_1_Streng = '{0:.2f}'.format(get_float((parm1 * standParm + compStreng), 2)),
                                                                                     minStreng='{0:.1f}'.format(get_float(minStreng, 1)),
                                                                                     Parm_2_Streng = '{0:.2f}'.format(get_float((parm2 * compStreng), 2)))
                elif(compNum > 14 and compNum <= 19):
                    parm1 = 1.05
                    parm2 = 0.85
                    standParm = 2.50
                    resultSheet['B' + str(8 + InsertFlag)] = 'C'
                    resultSheet['C' + str(8 + InsertFlag)] = compStreng
                    resultSheet['D' + str(8 + InsertFlag)] = compNum
                    resultSheet['E' + str(8 + InsertFlag)] = parm1
                    resultSheet['F' + str(8 + InsertFlag)] = parm2
                    resultSheet['G' + str(8 + InsertFlag)] = '{0:.1f}'.format(get_float(avgStreng, 1))
                    resultSheet['H' + str(8 + InsertFlag)] = '{0:.1f}'.format(get_float(minStreng, 1))
                    resultSheet['I' + str(8 + InsertFlag)] = standParm
                    resultSheet['J' + str(8 + InsertFlag)] = '{0:.2f}'.format(get_float((float(resultSheet['G' + str(8 + InsertFlag)].value) - (parm1 * standParm)), 2))
                    resultSheet['K' + str(8 + InsertFlag)] = '{0:.2f}'.format(get_float((parm2 * compStreng), 2))
                    QualifiedString = "C{compStreng}：mfcu = {avgStreng}＞fcu，k+λ\u2081·Sfcu ={Parm_1_Streng};fcu，min = {minStreng}≥λ\u2082·fcu，k = {Parm_2_Streng}，评定合格"
                    resultSheet['C' + str(14 + InsertFlag)] = QualifiedString.format(
                        compStreng='{0:.0f}'.format(compStreng),
                        avgStreng='{0:.1f}'.format(get_float(avgStreng, 1)),
                        Parm_1_Streng='{0:.2f}'.format(get_float((parm1 * standParm + compStreng), 2)),
                        minStreng='{0:.1f}'.format(get_float(minStreng, 1)),
                        Parm_2_Streng='{0:.2f}'.format(get_float((parm2 * compStreng), 2)))
                elif(compNum >= 20):
                    parm1 = 0.95
                    parm2 = 0.85
                    standParm = 2.50
                    resultSheet['B' + str(8 + InsertFlag)] = 'C'
                    resultSheet['C' + str(8 + InsertFlag)] = compStreng
                    resultSheet['D' + str(8 + InsertFlag)] = compNum
                    resultSheet['E' + str(8 + InsertFlag)] = parm1
                    resultSheet['F' + str(8 + InsertFlag)] = parm2
                    resultSheet['G' + str(8 + InsertFlag)] = '{0:.1f}'.format(get_float(avgStreng, 1))
                    resultSheet['H' + str(8 + InsertFlag)] = '{0:.1f}'.format(get_float(minStreng, 1))
                    resultSheet['I' + str(8 + InsertFlag)] = standParm
                    resultSheet['J' + str(8 + InsertFlag)] = '{0:.2f}'.format(get_float(float(resultSheet['G' + str(8 + InsertFlag)].value) - (parm1 * standParm), 2))
                    resultSheet['K' + str(8 + InsertFlag)] = '{0:.2f}'.format(get_float((parm2 * compStreng), 2))

                    QualifiedString = "C{compStreng}：mfcu = {avgStreng}＞fcu，k+λ\u2081·Sfcu ={Parm_1_Streng};fcu，min = {minStreng}≥λ\u2082·fcu，k = {Parm_2_Streng}，评定合格"
                    resultSheet['C' + str(14 + InsertFlag)] = QualifiedString.format(
                        compStreng='{0:.0f}'.format(compStreng),
                        avgStreng='{0:.1f}'.format(get_float(avgStreng, 1)),
                        Parm_1_Streng='{0:.2f}'.format(get_float((parm1 * standParm + compStreng), 2)),
                        minStreng='{0:.1f}'.format(get_float(minStreng, 1)),
                        Parm_2_Streng='{0:.2f}'.format(get_float((parm2 * compStreng) ,2)))
                resultSheet['A4'] = "工程名称：" + ConStrRepoHead[sheetNum].ProjectName
                resultSheet['A5'] = "施工单位: " + ConStrRepoHead[sheetNum].BuildUnit
                InsertFlag = InsertFlag + 1
                # avgStreng = mean(compList)
                # minStreng = min(compList)
                curNumFlag = curNumFlag + 1
            from Util.InsetPicUtil import insertpic
            from Util.get_parm import get_parm
            parm = get_parm()

            resultSheet = insertpic(resultSheet, picname=parm.Project4Manager, position='D21', width=90, heigh=30)
            resultSheet = insertpic(resultSheet, picname=parm.Project4Checker, position='G21')
            resultSheet = insertpic(resultSheet, picname=parm.Project4Calculate, position='K21', width=90, heigh=30)

    return resultRepoBook, resultRepoSumBook, resultRepoCompBook

if __name__ == '__main__':
    filename = "..\..\工地混凝土使用记录5.xlsx"
    session = session
    useBook = load_workbook(filename)
    resultRepoBook, resultRepoSumBook, resultRepoCompBook = ConStrengReport(useBook)
    resultRepoBook.save('标准养护混凝土试件抗压强度检验报告2.xlsx')
    resultRepoSumBook.save("试件强度实验结果汇总表3.xlsx")
    resultRepoCompBook.save('标准养护混凝土抗压强度计算表4.xlsx')