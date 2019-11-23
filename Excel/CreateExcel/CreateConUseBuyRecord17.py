# -*- coding:utf-8 -*-
"""
@author: Darcy
@time: 2018/10/28 16:10
"""
from openpyxl import load_workbook
from Util.ExcelInitUtil import iniexcel1, iniexcel7
from DateBase.connect_db import session
import Bean.ExcelBean
import Bean.DbBean
import Util.ExcelUtil
from DateBase.creat import CementAttributeDatum
from DateBase.creat import ConcreteMix
from collections import defaultdict
from Util.query_database import query_mix
from random import uniform
from datetime import timedelta, datetime
from Util.get_parm import get_parm
from Util.round import get_float
"""
    生成水泥购进，使用一览表
    Parm:
        useBook 使用记录表
        session 数据库连接池
    Return:
        modeBook 购进使用表
        CementDesignBook 配合比设计报告表
        maxRows 使用记录数量
"""
def ConUseBuyRecord(useBook):
    useSheets = useBook.sheetnames
    modeName = "../../Mode/1.xlsx"
    modeBook = load_workbook(modeName)
    """
        第一份表操作
    """
    """
        复制模板
    """
    for sheetNum in range(len(useSheets)):
        useSheet = useBook.worksheets[sheetNum]
        rows = useSheet.max_row - 9
        moreFixNum, rows = divmod(rows, 15)
        if rows == 0:
            moreFixNum = moreFixNum - 1
        for copyNum in range(moreFixNum + 1):
            modeSheet = modeBook.worksheets[0]
            iniexcel1(modeSheet)
            modeSheet = modeBook.copy_worksheet(modeSheet)
            iniexcel1(modeSheet)
            modeSheet.title = str(str(sheetNum + 1) + '-' + str(copyNum + 1))
    modeBook.remove(modeBook.worksheets[0])
    UseRecordListArr = []
    """
        获取数据
    """
    for sheetNum in range(len(useSheets)):
        sheet = useBook.worksheets[sheetNum]
        UseRecordList = []
        rows = sheet.max_row
        for rowNum in range(10, rows + 1):
            if sheet.cell(row=rowNum, column=1).value == None:
                break
            ProjectSite = sheet.cell(row=rowNum, column=1).value
            ConcreteName = sheet.cell(row=rowNum, column=2).value
            ConcreteStrength = sheet.cell(row=rowNum, column=3).value
            ImperLevel = sheet.cell(row=rowNum, column=4).value
            SwellLevel = sheet.cell(row=rowNum, column=5).value
            CuringDate = sheet.cell(row=rowNum, column=6).value
            CuringTime = sheet.cell(row=rowNum, column=7).value
            CuringNum = sheet.cell(row=rowNum, column=8).value
            ConUseRecord = Bean.DbBean.ConcreteUseRecord(
                ProjectSite=ProjectSite,
                ConcreteName=ConcreteName,
                ConcreteStrength=ConcreteStrength,
                ImperLevel=ImperLevel,
                SwellLevel=SwellLevel,
                CuringDate=CuringDate,
                CuringTime=CuringTime,
                CuringNum=CuringNum)
            UseRecordList.append(ConUseRecord)
        # UseRecordList.sort(key=lambda x: x.CuringDate, reverse=False)
        UseRecordListArr.append(UseRecordList)

    """
        获取使用记录表报告时间
    """
    repoTimeList = []
    for sheetNum in range(len(useSheets)):
        useSheet = useBook.worksheets[sheetNum]
        timeList = []
        rows = useSheet.max_row
        for useSheetRowNum in range(10, rows + 1):
            CuringDate = useSheet.cell(row=useSheetRowNum, column=6).value
            timeList.append(CuringDate)
            timeList.sort()
        repoTimeList.append(timeList[0])

    """
        插入数据
    """
    curSheetNum = 0
    CementAtrList = []
    repoTimeIndex = 0
    for sheetNum in range(len(useSheets)):
        sheet = useBook.worksheets[sheetNum]
        useRecordNum = len(UseRecordListArr[sheetNum])
        moreFixNum, rows = divmod(useRecordNum, 15)
        if rows == 0:
            moreFixNum = moreFixNum - 1
        CementAtrDataList = []
        for insertNum in range(moreFixNum + 1):
            InsertFlag = 0
            resultSheet = modeBook.worksheets[curSheetNum]
            ProjectName = sheet['A2'].value.replace('工程名称：', '').strip()
            ConstrUnit = sheet['A4'].value.replace('施工单位：', '').strip()
            resultSheet['A4'] = '工程名称：' + ProjectName
            resultSheet['N4'] = ConstrUnit
            curSheetNum = curSheetNum + 1
            for useNum in range(15):
                if((useNum + 15 * insertNum) == len(UseRecordListArr[sheetNum])):
                    if((useNum + 15 * insertNum) / 15 != 0):
                        resultSheet['M' + str(7 + InsertFlag)] = '以'
                        resultSheet['N' + str(7 + InsertFlag)] = '下'
                        resultSheet['O' + str(7 + InsertFlag)] = '空'
                        resultSheet['P' + str(7 + InsertFlag)] = '白'
                    break
                CementName = sheet['B' + str(10 + InsertFlag)].value.replace('砼','')
                UseRecord = UseRecordListArr[sheetNum][useNum + 15 * insertNum]
                query = session.query(CementAttributeDatum).order_by(
                    CementAttributeDatum.ArrivalTime.desc())
                result = query.filter(CementAttributeDatum.PriorityLevel == 1,
                                      CementAttributeDatum.ArrivalTime < UseRecord.CuringDate - timedelta(days=1)).all()
                CementAtr = result[0]
                resultSheet['A' + str(7 + InsertFlag)] = useNum + 15 * insertNum + 1
                resultSheet['B' + str(7 + InsertFlag)] = CementAtr.ArrivalTime.strftime('%Y-%#m-%#d')
                resultSheet['C' + str(7 + InsertFlag)] = CementAtr.CementVariety
                resultSheet['D' + str(7 + InsertFlag)] = CementAtr.Manufacturer
                resultSheet['E' + str(7 + InsertFlag)] = CementAtr.ProductionDate.strftime('%#m-%#d').replace('-','月') + '日'
                resultSheet['F' + str(7 + InsertFlag)] = CementAtr.CementId
                resultSheet['G' + str(7 + InsertFlag)] = CementAtr.CementNumber
                resultSheet['H' + str(7 + InsertFlag)] = Util.ExcelUtil.CemAtrIsStablility(CementAtr.IsStability)
                if(CementAtr.InitialTime[0:1] == '0'):
                    CementAtr.InitialTime = CementAtr.InitialTime[1:]
                if (CementAtr.FinalTime[0:1] == '0'):
                    CementAtr.FinalTime = CementAtr.FinalTime[1:]
                resultSheet['I' + str(7 + InsertFlag)] = CementAtr.InitialTime
                resultSheet['J' + str(7 + InsertFlag)] = CementAtr.FinalTime
                resultSheet['K' + str(7 + InsertFlag)] = CementAtr.R3_Compression
                resultSheet['L' + str(7 + InsertFlag)] = CementAtr.R28_Compression
                resultSheet['M' + str(7 + InsertFlag)] = UseRecord.ProjectSite
                # resultSheet['M' + str(7 + InsertFlag)] = Util.ExcelUtil.StringInLine(UseRecord.ProjectSite,11)

                StrengthString = UseRecord.ConcreteStrength
                if(UseRecord.ImperLevel == None):
                    StrengthString = StrengthString + CementName
                else:
                    StrengthString = StrengthString + CementName + UseRecord.ImperLevel
                if(UseRecord.SwellLevel != None):
                    StrengthString = StrengthString + '\n(' + str(int(UseRecord.SwellLevel * 100)) + '%膨胀)'
                resultSheet['N' + str(7 + InsertFlag)] = StrengthString
                resultSheet['O' + str(7 + InsertFlag)] = UseRecord.CuringDate.strftime('%Y-%#m-%#d')
                resultSheet['P' + str(7 + InsertFlag)] = UseRecord.CuringNum
                from Util.InsetPicUtil import insertpic
                parm = get_parm()
                resultSheet = insertpic(resultSheet, picname=parm.Project1Manager, position='C23')
                resultSheet = insertpic(resultSheet, picname=parm.Project1FillSheeter, position='G23', width=90, heigh=30)
                # 按照报告时间查询水泥
                QueryDate = repoTimeList[repoTimeIndex]
                query = session.query(CementAttributeDatum).order_by(
                    CementAttributeDatum.ArrivalTime.desc())
                result = query.filter(CementAttributeDatum.PriorityLevel == 1,
                                      CementAttributeDatum.ArrivalTime < QueryDate - timedelta(days=29)).all()
                CementAtr = result[0]
                conMixCementAtr = Bean.ExcelBean.ConMixCementAtr(ArrivalTime=CementAtr.ArrivalTime,
                                                                 CementId=CementAtr.CementId,
                                                                 CementName=UseRecord.ConcreteName,
                                                                 ConcreteStrength=UseRecord.ConcreteStrength,
                                                                 ImperLevel=Util.ExcelUtil.IsLevelNone(UseRecord.ImperLevel),
                                                                 SwellLevel=Util.ExcelUtil.IsLevelNone(UseRecord.SwellLevel),
                                                                 R3_Bending=CementAtr.R3_Bending,
                                                                 R28_Bending=CementAtr.R28_Bending,
                                                                 R3_Compression=CementAtr.R3_Compression,
                                                                 R28_Compression=CementAtr.R28_Compression)
                CementAtrDataList.append(conMixCementAtr)
                InsertFlag = InsertFlag + 1
        repoTimeIndex = repoTimeIndex + 1
        CementAtrList.append(CementAtrDataList)

    """
        第七份表操作
    """
    cementAtrBeanList = []
    for cementAtrBean in CementAtrList:
        cementAtrDataDict = defaultdict(list)
        for cementAtr in cementAtrBean:
            nameString = str(cementAtr.ConcreteStrength) + '$' + str(cementAtr.CementName) + '$' + str(cementAtr.ImperLevel) + '$' +  str(cementAtr.SwellLevel)
            cementAtrDataDict[nameString].append(cementAtr)
        cementAtrBeanList.append(cementAtrDataDict)

    modeName = "../../Mode/7.xlsx"
    CementDesignBook = load_workbook(modeName)
    useSheets = useBook.sheetnames
    UseRecordHeadList = []
    parm = get_parm()

    repoTimeIndex = 0
    for sheetNum in range(len(useSheets)):
        useSheet = useBook.worksheets[sheetNum]
        UseRecordList = []
        rows = useSheet.max_row
        """
            工地混凝土使用记录数据
        """
        ProjectName = useSheet['A2'].value.replace('工程名称：', '').strip()
        BuildUnit = useSheet['A3'].value.replace('建设单位：', '').strip()
        for useSheetRowNum in range(10, rows + 1):
            # CuringDate = useSheet.cell(row=useSheetRowNum, column=6).value
            CuringDate = repoTimeList[repoTimeIndex]
            for i in range(1, 40):
                time = str((CuringDate + timedelta(days=-i)).strftime('%#d'))
                if time == "1" or time == "10" or time == "20" or time == "30":
                    CuringDate = str((CuringDate + timedelta(days=-i)).strftime('%Y.%m.%d'))
                    break
            ConcreteName = useSheet.cell(row=useSheetRowNum, column=2).value
            ConcreteStrength = useSheet.cell(row=useSheetRowNum, column=3).value
            ImperLevel = useSheet.cell(row=useSheetRowNum, column=4).value
            SwellLevel = useSheet.cell(row=useSheetRowNum, column=5).value
            ConUseRecord = Bean.ExcelBean.ConMixUseBean(ProjectName=ProjectName,
                                                        BuildUnit=BuildUnit,
                                                        ConcreteName=ConcreteName,
                                                        ConcreteStrength=ConcreteStrength,
                                                        ImperLevel=Util.ExcelUtil.IsLevelNone(ImperLevel),
                                                        SwellLevel=Util.ExcelUtil.IsLevelNone(SwellLevel),
                                                        CuringDate=CuringDate,
                                                        )
            UseRecordList.append(ConUseRecord)
        repoTimeIndex = repoTimeIndex + 1
        UseRecordHeadList.append(UseRecordList)

    mixDesignBeanList = []
    for useRecordHeadList in UseRecordHeadList:
        mixDesignDataDict = defaultdict(list)
        for useRecord in useRecordHeadList:
            nameString = str(useRecord.ConcreteStrength) + '$' + str(useRecord.ConcreteName) + '$' + str(
                useRecord.ImperLevel) + '$' + str(useRecord.SwellLevel)
            mixDesignDataDict[nameString].append(useRecord)
        mixDesignBeanList.append(mixDesignDataDict)

    """
        复制模板
    """
    for sheetNum in range(len(useSheets)):
        rows = mixDesignBeanList[sheetNum].__len__()
        for insertNum in range(rows):
            modeSheet = CementDesignBook.worksheets[0]
            iniexcel7(modeSheet)
            modeSheet = CementDesignBook.copy_worksheet(modeSheet)
            iniexcel7(modeSheet)
            modeSheet.title = str(sheetNum + 1) + '-' + str(insertNum + 1)
    CementDesignBook.remove(CementDesignBook.worksheets[0])

    """
        插入数据
    """
    curSheetNum = 0
    for sheetNum in range(len(useSheets)):
        mixDesignDictKeys = list(mixDesignBeanList[sheetNum].keys())
        cenAtrDictKeys = list(cementAtrBeanList[sheetNum].keys())
        rows = mixDesignBeanList[sheetNum].__len__()
        for insertNum in range(rows):
            resultSheet = CementDesignBook.worksheets[curSheetNum]
            curSheetNum = curSheetNum + 1
            mixDesign = mixDesignBeanList[sheetNum][mixDesignDictKeys[insertNum]][0]
            cenAtr = cementAtrBeanList[sheetNum][mixDesignDictKeys[insertNum]][0]
            mixDesignNameList = mixDesignDictKeys[insertNum].split('$')
            ConcreteStrength = mixDesignNameList[0]
            ConcreteName = mixDesignNameList[1]
            ImperLevel = mixDesignNameList[2]
            SwellLevel = mixDesignNameList[3]
            if (SwellLevel != '/'):
                """
                    获取混合比数据
                """
                queryStreng = float(ConcreteStrength.strip('C'))
                queryName = ConcreteName
                if (ImperLevel == '/'):
                    ImperLevel = None
                queryImperLevel = ImperLevel
                querySwellLevel = float(float(SwellLevel.strip('%')))
                queryStreng = int(queryStreng)
                UseMix = query_mix(ConcreteName=queryName, StrengthLevel=queryStreng, ImperLevel=queryImperLevel, SwellLevel=querySwellLevel)
                # resultSheet['O3'] = "设计单位:" + parm[0].ConDesignUtil
                resultSheet['P4'] = "试验规格:" + parm.Project7ConDesignSpeciEdit
                """
                    膨胀表修改
                """
                resultSheet['M8'] = "膨胀"
                resultSheet['M9'] = '{0:.0f}'.format(querySwellLevel * 100) + '%'
                resultSheet['C17'] = "型号"
                resultSheet['H6'] = "粉煤灰"
                resultSheet['L16'] = "高效扛裂膨胀剂"
                resultSheet['L17'] = "HEA"
                resultSheet['P9'] = "泵送"
                # 接触掺合料合并
                resultSheet.unmerge_cells(start_row=19, end_row=19,
                                        start_column=13, end_column=14)
                resultSheet.unmerge_cells(start_row=20, end_row=20,
                                          start_column=13, end_column=14)
                resultSheet.unmerge_cells(start_row=24, end_row=24,
                                          start_column=13, end_column=14)
                resultSheet.unmerge_cells(start_row=25, end_row=25,
                                          start_column=13, end_column=14)
                # 设置字体格式
                from openpyxl.styles import Color, Font, Alignment
                font = Font(u'宋体', size=12, color='000000')
                resultSheet['N19'].font = font
                resultSheet['N24'].font = font
                resultSheet['N19'].alignment = Alignment(horizontal='center', vertical='center')
                resultSheet['N24'].alignment = Alignment(horizontal='center', vertical='center')
                resultSheet['N20'].font = font
                resultSheet['N25'].font = font
                resultSheet['N20'].alignment = Alignment(horizontal='center', vertical='center')
                resultSheet['N25'].alignment = Alignment(horizontal='center', vertical='center')

                resultSheet['N19'] = "膨胀剂"
                resultSheet['N24'] = "膨胀剂"

                """
                    插入工地混凝土使用记录
                """
                resultSheet['A3'] = "建设单位：" + mixDesign.BuildUnit
                resultSheet['A4'] = "工程名称：" + mixDesign.ProjectName
                resultSheet['A5'] = "发报告日期：" + mixDesign.CuringDate
                resultSheet['B9'] = mixDesign.ConcreteName
                resultSheet['H9'] = mixDesign.ConcreteStrength
                resultSheet['L9'] = mixDesign.ImperLevel
                """
                    插入配合比选用汇总表
                """
                if (UseMix.MixRatioName == None):
                    resultSheet['P5'] = "编号:"
                else:
                    resultSheet['P5'] = "编号:" + UseMix.MixRatioName
                resultSheet['O9'] = str(UseMix.SlumpNum).replace('', '')
                resultSheet['Q9'] = '{0:.1f}'.format(get_float(float(UseMix.StandardDeviation), 1))
                resultSheet['R9'] = '{0:.1f}'.format(get_float(float(UseMix.ConcreteStrengh), 1))
                resultSheet['P17'] = '{0:.1f}'.format(get_float(float(UseMix.AdmixtureAmount), 1))
                # resultSheet['B20'] = UseMix.CementRatio
                resultSheet['L20'] = UseMix.CementNum
                resultSheet['M20'] = UseMix.FlyashNum
                resultSheet['N20'] = int(UseMix.SwellingNum)
                resultSheet['O20'] = UseMix.SandNum
                resultSheet['P20'] = UseMix.GravelNum
                resultSheet['Q20'] = UseMix.WaterNum
                resultSheet['R20'] = '{0:.1f}'.format(get_float(float(UseMix.AdmixtureNum), 1))
                resultSheet['S20'] = Util.ExcelUtil.IsSwellingLevelNone(UseMix.SwellingNum)
                # resultSheet['H20'] = UseMix.SwellingNum
                resultSheet['B22'] = '{0:.1f}'.format(get_float((float(UseMix.SandRatio) * 100), 1)) + "%"
                resultSheet['C22'] = str(UseMix.SlumpNum).replace(' ', '')
                resultSheet['C27'] = str(UseMix.SlumpNum).replace(' ', '')


                """
                    插入水泥购进，使用情况一览表
                """
                resultSheet['M11'] = cenAtr.CementId
                resultSheet['O11'] = cenAtr.R3_Bending
                resultSheet['P11'] = cenAtr.R28_Bending
                resultSheet['Q11'] = cenAtr.R3_Compression
                resultSheet['R11'] = cenAtr.R28_Compression
                """
                           参数表调节数据插入
                       """
                compStrength = float(mixDesign.ConcreteStrength.strip('C'))
                resultSheet['L13'] = '{0:.1f}'.format(
                    uniform(float(parm.MinS_FinenessDensity), float(parm.MaxS_FinenessDensity)))
                resultSheet['M13'] = '{0:.0f}'.format(
                    uniform(float(parm.MinS_SurfaceDensity), float(parm.MaxS_SurfaceDensity)))
                resultSheet['O13'] = '{0:.0f}'.format(
                    uniform(float(parm.MinS_Density), float(parm.MaxS_Density)))
                resultSheet['Q13'] = uniform(float(parm.MinS_SlitContent), float(parm.MaxS_SlitContent))
                resultSheet['R13'] = '{0:.0f}'.format(
                    uniform(float(parm.MinS_WaterContent), float(parm.MaxS_WaterContent)))
                resultSheet['R13'] = uniform(float(parm.MinS_WaterContent), float(parm.MaxS_WaterContent))
                resultSheet['M15'] = '{0:.1f}'.format(
                    uniform(float(parm.MinG_GrainContent), float(parm.MaxG_GrainContent)))
                resultSheet['O15'] = uniform(float(parm.MinG_CrushLevel), float(parm.MaxG_CrushLevel))
                resultSheet['P15'] = '{0:.0f}'.format(
                    uniform(float(parm.MinG_Density), float(parm.MaxG_Density)))
                resultSheet['Q15'] = '{0:.1f}'.format(
                    uniform(float(parm.MinG_SlitContent), float(parm.MaxG_SlitContent)))
                # to do

                resultSheet['R15'] = '{0:.1f}'.format(
                    uniform(float(parm.MinG_WaterContent), float(parm.MaxG_WaterContent)))
                resultSheet['Q17'] = '{0:.1f}'.format(
                    uniform(float(parm.MinA_Density), float(parm.MaxA_Density)))
                resultSheet['M22'] = '{0:.1f}'.format(
                    compStrength * uniform(float(parm.MinR7_Compression), float(parm.MaxR7_Compression)))
                resultSheet['O22'] = '{0:.1f}'.format(
                    compStrength * uniform(float(parm.MinR28_Compression), float(parm.MaxR28_Compression)))
                resultSheet['M27'] = '{0:.1f}'.format(
                    compStrength * uniform(float(parm.MinR7_Compression), float(parm.MaxR7_Compression)))
                resultSheet['O27'] = '{0:.1f}'.format(
                    compStrength * uniform(float(parm.MinR28_Compression), float(parm.MaxR28_Compression)))
                """
                    质量比公式数据插入
                """
                # 干料
                resultSheet['C20'] = 1
                # resultSheet['E20'] = float(resultSheet['M20'].value) / float(resultSheet['L20'].value)
                # resultSheet['G20'] = float(resultSheet['N20'].value) / float(resultSheet['L20'].value)
                # resultSheet['I20'] = float(resultSheet['O20'].value) / float(resultSheet['L20'].value)
                # resultSheet['K20'] = float(resultSheet['P20'].value) / float(resultSheet['L20'].value)
                # resultSheet['K20'] = resultSheet['Q20'].value / resultSheet['L20'].value
                admixtureNum = float(resultSheet['M20'].value) / float(resultSheet['L20'].value)
                imperviousNum = float(resultSheet['N20'].value) / float(resultSheet['L20'].value)
                sandNum = float(resultSheet['O20'].value) / float(resultSheet['L20'].value)
                stoneNum = float(resultSheet['P20'].value) / float(resultSheet['L20'].value)
                waterNum = float(resultSheet['Q20'].value) / float(resultSheet['L20'].value)
                QualifiedString = "  1：{admixtureNum} ：{imperviousNum} ：{sandNum} ：{stoneNum} ：{waterNum}"
                resultSheet["C20"] = QualifiedString.format(
                    admixtureNum = '{0:.2f}'.format(admixtureNum),
                    imperviousNum = '{0:.2f}'.format(imperviousNum),
                    sandNum = '{0:.2f}'.format(sandNum),
                    stoneNum = '{0:.2f}'.format(stoneNum),
                    waterNum = '{0:.2f}'.format(waterNum),
                )
                # 施工
                resultSheet['L25'] = resultSheet['L20'].value
                resultSheet['M25'] = resultSheet['M20'].value
                # 3.26
                resultSheet['N25'] = resultSheet['N20'].value
                resultSheet['O25'] = '{0:.0f}'.format(
                    float(resultSheet['O20'].value) + float(resultSheet['O20'].value) * resultSheet['R13'].value)
                resultSheet['P25'] = '{0:.0f}'.format(
                    float(resultSheet['P20'].value) + float(resultSheet['P20'].value) * (
                            float(resultSheet['R15'].value) / 100))
                # resultSheet['Q25'] = '{0:.0f}'.format(
                #     float(resultSheet['Q20'].value) - float(resultSheet['O20'].value) * resultSheet[
                #         'R13'].value - float(resultSheet['P20'].value) * (float(resultSheet['R15'].value) / 100))
                resultSheet['Q25'] = '{0:.0f}'.format(
                    float(resultSheet['Q20'].value) -
                    (
                        float(resultSheet['O25'].value) + float(resultSheet["P25"].value) - float(resultSheet['O20'].value) - float(resultSheet['P20'].value)
                    )
                )
                resultSheet['R25'] = resultSheet['R20'].value
                resultSheet['S25'] = resultSheet['S20'].value
                # resultSheet['C25'] = 1
                # resultSheet['E25'] = float(resultSheet['M25'].value) / float(resultSheet['L25'].value)
                # resultSheet['G25'] = '{0:.2f}'.format(float(resultSheet['N25'].value) / float(resultSheet['L25'].value))
                # resultSheet['I25'] = '{0:.2f}'.format(float(resultSheet['O25'].value) / float(resultSheet['L25'].value))
                # resultSheet['K25'] = '{0:.2f}'.format(float(resultSheet['P25'].value) / float(resultSheet['L25'].value))
                # resultSheet['K25'] = '{0:.2f}'.format(float(resultSheet['Q25'].value) / float(resultSheet['L25'].value))
                admixtureNum = float(resultSheet['M25'].value) / float(resultSheet['L25'].value)
                imperviousNum = float(resultSheet['N25'].value) / float(resultSheet['L25'].value)
                sandNum = float(resultSheet['O25'].value) / float(resultSheet['L25'].value)
                stoneNum = float(resultSheet['P25'].value) / float(resultSheet['L25'].value)
                waterNum = float(resultSheet['Q25'].value) / float(resultSheet['L25'].value)
                QualifiedString = "  1：{admixtureNum} ：{imperviousNum} ：{sandNum} ：{stoneNum} ：{waterNum}"
                resultSheet["C25"] = QualifiedString.format(
                    admixtureNum='{0:.2f}'.format(get_float(float(admixtureNum), 2)),
                    imperviousNum='{0:.2f}'.format(get_float(float(imperviousNum), 2)),
                    sandNum='{0:.2f}'.format(get_float(float(sandNum), 2)),
                    stoneNum='{0:.2f}'.format(get_float(float(stoneNum), 2)),
                    waterNum='{0:.2f}'.format(get_float(float(waterNum), 2)),
                )
                resultSheet['B20'] = '{0:.2f}'.format(float(resultSheet['Q20'].value) / (
                        float(resultSheet['L20'].value) + float(resultSheet['M20'].value) + float(resultSheet['N20'].value)))
                resultSheet['B25'] = '{0:.2f}'.format(float(resultSheet['Q25'].value) / (
                        float(resultSheet['L25'].value) + float(resultSheet['M25'].value) + float(resultSheet['N25'].value)))

                resultSheet['B27'] = '{0:.1f}'.format(float(resultSheet['O25'].value) / (
                        float(resultSheet['O25'].value) + float(resultSheet['P25'].value)) * 100) + '%'
                resultSheet['R13'] = '{0:.1f}'.format(resultSheet['R13'].value * 100)
                resultSheet['H22'] = '{0:.0f}'.format(get_float(float(resultSheet["L20"].value) + float(resultSheet["M20"].value) + float(resultSheet["N20"].value) + float(resultSheet["O20"].value) + float(resultSheet["P20"].value) + float(resultSheet["Q20"].value) + float(resultSheet["R20"].value), 0))
                resultSheet['H27'] = '{0:.0f}'.format(get_float(float(resultSheet["L25"].value) + float(resultSheet["M25"].value) + float(resultSheet["N25"].value) + float(resultSheet["O25"].value) + float(resultSheet["P25"].value) + float(resultSheet["Q25"].value) + float(resultSheet["R25"].value), 0))
            else:
                """
                    获取混合比数据
                """
                queryStreng = float(ConcreteStrength.strip('C'))
                queryName = ConcreteName
                if(queryName == "泵送砼"):
                    resultSheet['O9'] = "泵送"
                if (ImperLevel == '/'):
                    ImperLevel = None
                if (SwellLevel == '/'):
                    SwellLevel = None
                queryImperLevel = ImperLevel
                querySwellLevel = SwellLevel
                queryStreng = int(queryStreng)
                UseMix = query_mix(ConcreteName=queryName, StrengthLevel=queryStreng, ImperLevel=queryImperLevel, SwellLevel=querySwellLevel)
                # resultSheet['O3'] = "设计单位:" + parm[0].ConDesignUtil
                resultSheet['P4'] = "试验规格:" + parm.Project7ConDesignSpeciEdit
                """
                    插入工地混凝土使用记录
                """
                resultSheet['A3'] = "建设单位：" + mixDesign.BuildUnit
                resultSheet['A4'] = "工程名称：" + mixDesign.ProjectName
                resultSheet['A5'] = "发报告日期：" + mixDesign.CuringDate
                resultSheet['B9'] = mixDesign.ConcreteName
                resultSheet['H9'] = mixDesign.ConcreteStrength
                resultSheet['L9'] = mixDesign.ImperLevel
                """
                    插入配合比选用汇总表
                """
                if (UseMix.MixRatioName == None):
                    resultSheet['P5'] = "编号:"
                else:
                    resultSheet['P5'] = "编号:" + UseMix.MixRatioName
                resultSheet['O9'] = UseMix.SlumpNum
                resultSheet['Q9'] = '{0:.1f}'.format(get_float(float(UseMix.StandardDeviation), 1))
                resultSheet['R9'] = '{0:.1f}'.format(get_float(float(UseMix.ConcreteStrengh), 1))
                resultSheet['P17'] = '{0:.1f}'.format(get_float(float(UseMix.AdmixtureAmount), 1))
                # resultSheet['B20'] = UseMix.CementRatio
                resultSheet['L20'] = UseMix.CementNum
                resultSheet['M20'] = UseMix.FlyashNum
                resultSheet['O20'] = UseMix.SandNum
                resultSheet['P20'] = UseMix.GravelNum
                resultSheet['Q20'] = UseMix.WaterNum
                resultSheet['R20'] = '{0:.1f}'.format(get_float(float(UseMix.AdmixtureNum), 1))
                resultSheet['S20'] = Util.ExcelUtil.IsSwellingLevelNone(UseMix.SwellingNum)
                # resultSheet['H20'] = UseMix.SwellingNum
                resultSheet['B22'] = '{0:.1f}'.format(get_float((float(UseMix.SandRatio) * 100), 1)) + "%"
                resultSheet['C22'] = UseMix.SlumpNum
                resultSheet['C27'] = UseMix.SlumpNum
                # resultSheet['H22'] = UseMix.MassDensity
                # resultSheet['H27'] = UseMix.MassDensity
                """
                    插入水泥购进，使用情况一览表
                """
                resultSheet['M11'] = cenAtr.CementId
                resultSheet['O11'] = cenAtr.R3_Bending
                resultSheet['P11'] = cenAtr.R28_Bending
                resultSheet['Q11'] = cenAtr.R3_Compression
                resultSheet['R11'] = cenAtr.R28_Compression
                """
                           参数表调节数据插入
                       """
                compStrength = float(mixDesign.ConcreteStrength.strip('C'))
                resultSheet['L13'] = '{0:.1f}'.format(
                    uniform(float(parm.MinS_FinenessDensity), float(parm.MaxS_FinenessDensity)))
                resultSheet['M13'] = '{0:.0f}'.format(
                    uniform(float(parm.MinS_SurfaceDensity), float(parm.MaxS_SurfaceDensity)))
                resultSheet['O13'] = '{0:.0f}'.format(
                    uniform(float(parm.MinS_Density), float(parm.MaxS_Density)))
                resultSheet['Q13'] = uniform(float(parm.MinS_SlitContent), float(parm.MaxS_SlitContent))
                resultSheet['P13'] = '{0:.1f}'.format(
                    uniform(float(parm.MinS_WaterContent), float(parm.MaxS_WaterContent)))
                resultSheet['R13'] = uniform(float(parm.MinS_WaterContent), float(parm.MaxS_WaterContent))
                resultSheet['M15'] = '{0:.1f}'.format(
                    uniform(float(parm.MinG_GrainContent), float(parm.MaxG_GrainContent)))
                resultSheet['O15'] = uniform(float(parm.MinG_CrushLevel), float(parm.MaxG_CrushLevel))
                resultSheet['P15'] = '{0:.0f}'.format(
                    uniform(float(parm.MinG_Density), float(parm.MaxG_Density)))
                resultSheet['Q15'] = '{0:.1f}'.format(
                    uniform(float(parm.MinG_SlitContent), float(parm.MaxG_SlitContent)))
                # to do
                resultSheet['R15'] = '{0:.1f}'.format(
                    uniform(float(parm.MinG_WaterContent), float(parm.MaxG_WaterContent)))
                resultSheet['Q17'] = '{0:.1f}'.format(
                    uniform(float(parm.MinA_Density), float(parm.MaxA_Density)))
                resultSheet['M22'] = '{0:.1f}'.format(
                    compStrength * uniform(float(parm.MinR7_Compression), float(parm.MaxR7_Compression)))
                resultSheet['O22'] = '{0:.1f}'.format(
                    compStrength * uniform(float(parm.MinR28_Compression), float(parm.MaxR28_Compression)))
                resultSheet['M27'] = '{0:.1f}'.format(
                    compStrength * uniform(float(parm.MinR7_Compression), float(parm.MaxR7_Compression)))
                resultSheet['O27'] = '{0:.1f}'.format(
                    compStrength * uniform(float(parm.MinR28_Compression), float(parm.MaxR28_Compression)))
                """
                    质量比公式数据插入
                """
                # 干料
                # resultSheet['C20'] = 1
                # resultSheet['E20'] = float(resultSheet['M20'].value) / float(resultSheet['L20'].value)
                # resultSheet['G20'] = float(resultSheet['O20'].value) / float(resultSheet['L20'].value)
                # resultSheet['I20'] = float(resultSheet['P20'].value) / float(resultSheet['L20'].value)
                # resultSheet['K20'] = float(resultSheet['Q20'].value) / float(resultSheet['L20'].value)
                admixtureNum = float(resultSheet['M20'].value) / float(resultSheet['L20'].value)
                # imperviousNum = float(resultSheet['N20'].value) / float(resultSheet['L20'].value)
                sandNum = float(resultSheet['O20'].value) / float(resultSheet['L20'].value)
                stoneNum = float(resultSheet['P20'].value) / float(resultSheet['L20'].value)
                waterNum = float(resultSheet['Q20'].value) / float(resultSheet['L20'].value)
                QualifiedString = "  1 ：{admixtureNum} ：{sandNum} ：{stoneNum} ：{waterNum}  "
                resultSheet["C20"] = QualifiedString.format(
                    admixtureNum='{0:.2f}'.format(get_float(float(admixtureNum), 2)),
                    sandNum='{0:.2f}'.format(get_float(float(sandNum), 2)),
                    stoneNum='{0:.2f}'.format(get_float(float(stoneNum), 2)),
                    waterNum='{0:.2f}'.format(get_float(float(waterNum), 2)),
                )
                # 施工
                resultSheet['L25'] = resultSheet['L20'].value
                resultSheet['M25'] = resultSheet['M20'].value
                resultSheet['O25'] = '{0:.0f}'.format(
                    float(resultSheet['O20'].value) + float(resultSheet['O20'].value) * resultSheet['R13'].value)
                resultSheet['P25'] = '{0:.0f}'.format(
                    float(resultSheet['P20'].value) + float(resultSheet['P20'].value) * (
                            float(resultSheet['R15'].value) / 100))
                # resultSheet['Q25'] = '{0:.0f}'.format(
                #     float(resultSheet['Q20'].value) - float(resultSheet['O20'].value) * resultSheet[
                #         'R13'].value - float(resultSheet['P20'].value) * (float(resultSheet['R15'].value) / 100))
                resultSheet['Q25'] = '{0:.0f}'.format(
                    float(resultSheet['Q20'].value) -
                    (
                            float(resultSheet['O25'].value) + float(resultSheet["P25"].value) - float(
                        resultSheet['O20'].value) - float(resultSheet['P20'].value)
                    )
                )

                resultSheet['R25'] = resultSheet['R20'].value
                resultSheet['S25'] = resultSheet['S20'].value
                # resultSheet['C25'] = 1
                # resultSheet['E25'] = float(resultSheet['M25'].value) / float(resultSheet['L25'].value)
                # resultSheet['G25'] = '{0:.2f}'.format(float(resultSheet['O25'].value) / float(resultSheet['L25'].value))
                # resultSheet['I25'] = '{0:.2f}'.format(float(resultSheet['P25'].value) / float(resultSheet['L25'].value))
                # resultSheet['K25'] = '{0:.2f}'.format(float(resultSheet['Q25'].value) / float(resultSheet['L25'].value))
                admixtureNum = float(resultSheet['M25'].value) / float(resultSheet['L25'].value)
                # imperviousNum = float(resultSheet['N25'].value) / float(resultSheet['L25'].value)
                sandNum = float(resultSheet['O25'].value) / float(resultSheet['L25'].value)
                stoneNum = float(resultSheet['P25'].value) / float(resultSheet['L25'].value)
                waterNum = float(resultSheet['Q25'].value) / float(resultSheet['L25'].value)
                QualifiedString = "  1 ：{admixtureNum} ：{sandNum} ：{stoneNum} ：{waterNum}  "
                resultSheet["C25"] = QualifiedString.format(
                    admixtureNum='{0:.2f}'.format(get_float(float(admixtureNum), 2)),
                    # imperviousNum='{0:.2f}'.format(imperviousNum),
                    sandNum='{0:.2f}'.format(get_float(float(sandNum), 2)),
                    stoneNum='{0:.2f}'.format(get_float(float(stoneNum), 2)),
                    waterNum='{0:.2f}'.format(get_float(float(waterNum), 2)),
                )
                resultSheet['B20'] = '{0:.2f}'.format(float(resultSheet['Q20'].value) / (
                        float(resultSheet['L20'].value) + float(resultSheet['M20'].value)))
                resultSheet['B25'] = '{0:.2f}'.format(float(resultSheet['Q25'].value) / (
                            float(resultSheet['L25'].value) + float(resultSheet['M25'].value)))
                resultSheet['B27'] = '{0:.1f}'.format((float(resultSheet['O25'].value) / (
                            float(resultSheet['O25'].value) + float(resultSheet['P25'].value)) * 100)) + '%'
                resultSheet['R13'] = '{0:.1f}'.format(resultSheet['R13'].value * 100)
                resultSheet['H22'] = '{0:.0f}'.format(get_float((float(resultSheet["L20"].value) + float(resultSheet["M20"].value) + float(resultSheet["O20"].value) + float(resultSheet[
                    "P20"].value) + float(resultSheet["Q20"].value) + float(resultSheet["R20"].value)), 0))
                resultSheet['H27'] = '{0:.0f}'.format(get_float((
                        float(resultSheet["L25"].value) + float(resultSheet["M25"].value) + float(resultSheet["O25"].value) + float(resultSheet[
                    "P25"].value) + float(resultSheet["Q25"].value) + float(resultSheet["R25"].value)), 0))
            from Util.InsetPicUtil import insertpic
            resultSheet = insertpic(resultSheet, picname=parm.Project7Manager, position='B29', width=90, heigh=30)
            resultSheet = insertpic(resultSheet, picname=parm.Project7Checker, position='M29')
            resultSheet = insertpic(resultSheet, picname=parm.Project7try, position='Q29', width=80, heigh=35)

    #  删除第七份表重复的
    useSheets = CementDesignBook.sheetnames
    num = []    # 记录要删除的号码
    for i in range(len(useSheets)):
        useSheet1 = CementDesignBook.worksheets[i]
        name1 = useSheet1['B9'].value
        strength1 = useSheet1['H9'].value
        permeability1 = useSheet1['L9'].value
        swell1 = useSheet1['M9'].value
        projectName1 = useSheet1['A4'].value
        for j in range(i):
            useSheet2 = CementDesignBook.worksheets[j]
            name2 = useSheet2['B9'].value
            strength2 = useSheet2['H9'].value
            permeability2 = useSheet2['L9'].value
            swell2 = useSheet2['M9'].value
            projectName2 = useSheet2['A4'].value
            if(projectName1 == projectName2):
                if (name1 == name2 and strength1 == strength2 and permeability1==permeability2 and swell1==swell2):
                    num.append(i)
                    break

    for i in num:
        CementDesignBook.remove(CementDesignBook.worksheets[i])
        for j in range(len(num)):
            num[j] = num[j]-1
    return modeBook, CementDesignBook

if __name__ == '__main__':
    filename = "..\..\工地混凝土使用记录8.xlsx"
    useBook = load_workbook(filename)
    session = session
    result, MixDesignBook = ConUseBuyRecord(useBook)
    result.save('水泥购进，使用一览表1.xlsx')
    MixDesignBook.save("配合比设计报告表7.xlsx")
    pass