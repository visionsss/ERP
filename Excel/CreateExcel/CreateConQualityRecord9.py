# coding=utf-8
"""
Author: vision
date: 2019/4/7 15:04
"""
from openpyxl import load_workbook
from Util.get_parm import get_parm
from DateBase.creat import ConcreteUseRecord
from Util.ExcelInitUtil import iniexcel
from datetime import timedelta
from random import uniform
from Util.InsetPicUtil import insertpic
class Sheet():
    def __init__(self):
        self.B4 = ''
        self.G4 = ''
        self.B6 = ''
        self.B7 = ''
        # self.C7 = ''
        self.G7 = ''
        self.H7 = ''
        self.I7 = ''
        self.J7 = ''
        self.K7 = ''
        self.L7 = ''
        self.M7 = ''
        self.G8 = ''
        self.K8 = ''
        self.K9 = ''
        self.K10 = ''
        self.C28 = ''
        self.I27 = 'R1'
        self.I28 = ''
        self.I29 = ''
        self.I30 = ''
        self.J27 = 'R3'
        self.J28 = ''
        self.J29 = ''
        self.J30 = ''
        self.K27 = 'R7'
        self.K28 = ''
        self.K29 = ''
        self.K30 = ''
        self.L28 = ''
        self.L29 = ''
        self.L30 = ''
def findR28(i, j, book1):
    i = str(i)
    j = str(j)
    for ii in book1.sheetnames:
        num = ii.split('-')[0]
        if num == str(i):
            for row in range(7, 27):
                if str(book1[ii].cell(row=row, column=1).value) == j:
                    s = book1[ii].cell(row=row, column=6).value
                    s = str(s)
                    s = s.split(r"/")

                    return s

'''
找配合比所需的数据
'''
def findmix(i, ConcreteName, ConcreteStrength, ImperLevel, SwellLevel, book2):
    if ImperLevel == None:
        ImperLevel = '/'
    if SwellLevel == None:
        SwellLevel = '/'
    else:
        SwellLevel *= 100
        SwellLevel = str(int(SwellLevel)) + '%'
    for ii in book2.sheetnames:
        if (str(book2[ii].cell(row=9, column=2).value) == str(ConcreteName) and
            str(book2[ii].cell(row=9, column=8).value) == str(ConcreteStrength) and
            str(book2[ii].cell(row=9, column=12).value) == str(ImperLevel) and
            str(book2[ii].cell(row=9, column=13).value) == str(SwellLevel)):
                WaterNum = book2[ii].cell(row=25, column=17).value
                CementNum = book2[ii].cell(row=25, column=12).value
                SandNum = book2[ii].cell(row=25, column=15).value
                GravelNum = book2[ii].cell(row=25, column=16).value
                AdmixtureAmount = book2[ii].cell(row=25, column=13).value
                AdmixtureNum = book2[ii].cell(row=25, column=18).value
                SlumpNum = str(book2[ii].cell(row=27, column=3).value).replace(' ', '')
                swell = -1
                if book2[ii].cell(row=19, column=14).value == '膨胀剂':
                    swell = book2[ii].cell(row=20, column=14).value
                return WaterNum, CementNum, SandNum, GravelNum, AdmixtureAmount, AdmixtureNum, SlumpNum, swell

def ConQualityRecord(useBook, book1, book2):
    modeName = "../../Mode/9.xlsx"
    modebook = load_workbook(modeName)
    useSheets = useBook.sheetnames
    UseRecordListArr = []
    """
        获取数据
    """
    ProjectNames = []
    ProjectUnits = []
    for sheetNum in range(len(useSheets)):
        sheet = useBook.worksheets[sheetNum]
        ProjectNames.append(sheet.cell(row=2, column=1).value)
        ProjectUnits.append(sheet.cell(row=4, column=1).value)
        UseRecordList = []
        rows = sheet.max_row
        for rowNum in range(10, rows + 1):
            if sheet.cell(row=rowNum, column=1).value == None:
                break
            ProjectSite = sheet.cell(row=rowNum, column=1).value
            ConcreteName = sheet.cell(row=rowNum, column=2).value
            ConcreteStrength = sheet.cell(row=rowNum, column=3).value
            ImperLevel = sheet.cell(row=rowNum, column=4).value
            SwellLevel = sheet.cell(row=rowNum, column=5).value
            CuringDate = sheet.cell(row=rowNum, column=6).value
            CuringTime = sheet.cell(row=rowNum, column=7).value
            CuringNum = sheet.cell(row=rowNum, column=8).value
            ConUseRecord = ConcreteUseRecord(
                ProjectSite=ProjectSite,
                ConcreteName=ConcreteName,
                ConcreteStrength=ConcreteStrength,
                ImperLevel=ImperLevel,
                SwellLevel=SwellLevel,
                CuringDate=CuringDate,
                CuringTime=CuringTime,
                CuringNum=CuringNum)
            UseRecordList.append(ConUseRecord)
        UseRecordListArr.append(UseRecordList)
    '''
       填数据
       '''

    curSheetNum = 0
    parm = get_parm()
    MinS_SlitContent = float(parm.MinS_SlitContent)
    MaxS_SlitContent = float(parm.MaxS_SlitContent)
    MinG_SlitContent = float(parm.MinG_SlitContent)
    MaxG_SlitContent = float(parm.MaxG_SlitContent)
    for sheetNum in range(len(useSheets)):
        useSheet = useBook.worksheets[sheetNum]
        rows = useSheet.max_row
        ProjectName = ProjectNames[sheetNum].replace('工程名称：', '')
        ProjectUnit = ProjectUnits[sheetNum].replace('施工单位：', '')
        count = 0
        for rowNum in range(10, rows + 1):
            useMessage = UseRecordListArr[sheetNum][count]
            count += 1
            R28 = findR28(sheetNum + 1, rowNum - 9, book1)
            R28len = len(R28)
            # print(R28, sheetNum+1, rowNum - 9)
            for iii in range((len(R28) + 11) // 12):
                curSheet = modebook.copy_worksheet(modebook.worksheets[0])
                iniexcel(curSheet, 9)
                curSheet.title = str(sheetNum + 1) + '#' + str(rowNum - 9)
                fill = Sheet()
                fill.B4 = ProjectName
                fill.G4 = '施工单位：' + ProjectUnit
                fill.B6 = useMessage.ProjectSite
                date = str(useMessage.CuringDate.strftime('%Y/%#m/%#d')).split("/")
                t = str(useMessage.CuringTime.replace('~','-').replace('_', ''))
                t = t.split('-')
                # print(t)
                if len(t[0]) == 4:
                    t[0] = '0' + t[0]
                if len(t[1]) == 4:
                    t[1] = '0' + t[1]
                fill.B7 = date[0] + "年" + date[1] + "月" + date[2] + "日" + t[0] + '-' + t[1]
                # fill.C7 = useMessage.CuringTime
                # print(sheetNum + 1,useMessage.ConcreteName,useMessage.ConcreteStrength,useMessage.ImperLevel,useMessage.SwellLevel)
                # aa = findmix(i=sheetNum + 1, ConcreteName=useMessage.ConcreteName,
                #         ConcreteStrength=useMessage.ConcreteStrength,
                #         ImperLevel=useMessage.ImperLevel, SwellLevel=useMessage.SwellLevel, book2=book2)
                # print(aa)
                fill.G7, fill.H7, fill.I7, fill.J7, fill.K7, fill.M7, fill.C28, fill.L7 = \
                    findmix(i=sheetNum + 1, ConcreteName=useMessage.ConcreteName,
                            ConcreteStrength=useMessage.ConcreteStrength,
                            ImperLevel=useMessage.ImperLevel, SwellLevel=useMessage.SwellLevel, book2=book2)
                for i in range(1, 40):
                    time = str((useMessage.CuringDate + timedelta(days=-i)).strftime('%#d'))

                    if time == "1" or time == "10" or time == "20" or time == "30":
                        fill.G8 = str((useMessage.CuringDate + timedelta(days=-i)).strftime('%Y/%#m/%#d')).replace('/',
                                                                                                                   '-')
                        # print("G8:  ",fill.G8)
                        break

                # fill.G8 = str((useMessage.CuringDate + timedelta(days=-2)).strftime('%Y/%#m/%#d')).replace('/', '-')
                ImperLevel = ''
                SwellLevel = ''
                if useMessage.ImperLevel == None:
                    ImperLevel = ''
                else:
                    ImperLevel = useMessage.ImperLevel
                if useMessage.SwellLevel == None:
                    SwellLevel = ''
                else:
                    SwellLevel = useMessage.SwellLevel * 100
                    SwellLevel = '\n（' + str(int(SwellLevel)) + "%膨胀）"
                fill.K8 = str(useMessage.ConcreteStrength) + str(useMessage.ConcreteName.replace('砼', '')) + str(
                    ImperLevel) + str(SwellLevel)
                fill.K9 = str(float('%.1f' % uniform(float(MinS_SlitContent), float(MaxS_SlitContent)))) + '%'
                fill.K10 = str(float('%.1f' % uniform(float(MinG_SlitContent), float(MaxG_SlitContent)))) + '%'
                # print(R28len,iii)
                if R28len >= 12:
                    ccc = 0
                else:
                    ccc = R28len % 12
                if ccc == 1:
                    fill.L28 = R28[iii * 12 + 0]
                elif ccc == 2:
                    fill.L28 = R28[iii * 12 + 0]
                    fill.L29 = R28[iii * 12 + 1]
                elif ccc == 3:
                    fill.L28 = R28[iii * 12 + 0]
                    fill.L29 = R28[iii * 12 + 1]
                    fill.L30 = R28[iii * 12 + 2]
                elif ccc == 4:
                    fill.L28 = R28[iii * 12 + 0]
                    fill.L29 = R28[iii * 12 + 1]
                    fill.L30 = R28[iii * 12 + 2]
                    fill.K27 = 'R28'
                    fill.K28 = R28[iii * 12 + 3]
                elif ccc == 5:
                    fill.L28 = R28[iii * 12 + 0]
                    fill.L29 = R28[iii * 12 + 1]
                    fill.L30 = R28[iii * 12 + 2]
                    fill.K27 = 'R28'
                    fill.K28 = R28[iii * 12 + 3]
                    fill.K29 = R28[iii * 12 + 4]
                elif ccc == 6:
                    fill.L28 = R28[iii * 12 + 0]
                    fill.L29 = R28[iii * 12 + 1]
                    fill.L30 = R28[iii * 12 + 2]
                    fill.K27 = 'R28'
                    fill.K28 = R28[iii * 12 + 3]
                    fill.K29 = R28[iii * 12 + 4]
                    fill.K30 = R28[iii * 12 + 5]
                elif ccc == 7:
                    fill.L28 = R28[iii * 12 + 0]
                    fill.L29 = R28[iii * 12 + 1]
                    fill.L30 = R28[iii * 12 + 2]
                    fill.K27 = 'R28'
                    fill.K28 = R28[iii * 12 + 3]
                    fill.K29 = R28[iii * 12 + 4]
                    fill.K30 = R28[iii * 12 + 5]
                    fill.J27 = 'R28'
                    fill.J28 = R28[iii * 12 + 6]
                elif ccc == 8:
                    fill.L28 = R28[iii * 12 + 0]
                    fill.L29 = R28[iii * 12 + 1]
                    fill.L30 = R28[iii * 12 + 2]
                    fill.K27 = 'R28'
                    fill.K28 = R28[iii * 12 + 3]
                    fill.K29 = R28[iii * 12 + 4]
                    fill.K30 = R28[iii * 12 + 5]
                    fill.J27 = 'R28'
                    fill.J28 = R28[iii * 12 + 6]
                    fill.J29 = R28[iii * 12 + 7]
                elif ccc == 9:
                    fill.L28 = R28[iii * 12 + 0]
                    fill.L29 = R28[iii * 12 + 1]
                    fill.L30 = R28[iii * 12 + 2]
                    fill.K27 = 'R28'
                    fill.K28 = R28[iii * 12 + 3]
                    fill.K29 = R28[iii * 12 + 4]
                    fill.K30 = R28[iii * 12 + 5]
                    fill.J27 = 'R28'
                    fill.J28 = R28[iii * 12 + 6]
                    fill.J29 = R28[iii * 12 + 7]
                    fill.J30 = R28[iii * 12 + 8]
                elif ccc == 10:
                    fill.L28 = R28[iii * 12 + 0]
                    fill.L29 = R28[iii * 12 + 1]
                    fill.L30 = R28[iii * 12 + 2]
                    fill.K27 = 'R28'
                    fill.K28 = R28[iii * 12 + 3]
                    fill.K29 = R28[iii * 12 + 4]
                    fill.K30 = R28[iii * 12 + 5]
                    fill.J27 = 'R28'
                    fill.J28 = R28[iii * 12 + 6]
                    fill.J29 = R28[iii * 12 + 7]
                    fill.J30 = R28[iii * 12 + 8]
                    fill.I27 = 'R28'
                    fill.I28 = R28[iii * 12 + 9]
                elif ccc == 11:
                    fill.L28 = R28[iii * 12 + 0]
                    fill.L29 = R28[iii * 12 + 1]
                    fill.L30 = R28[iii * 12 + 2]
                    fill.K27 = 'R28'
                    fill.K28 = R28[iii * 12 + 3]
                    fill.K29 = R28[iii * 12 + 4]
                    fill.K30 = R28[iii * 12 + 5]
                    fill.J27 = 'R28'
                    fill.J28 = R28[iii * 12 + 6]
                    fill.J29 = R28[iii * 12 + 7]
                    fill.J30 = R28[iii * 12 + 8]
                    fill.I27 = 'R28'
                    fill.I28 = R28[iii * 12 + 9]
                    fill.I29 = R28[iii * 12 + 10]
                elif ccc == 0:
                    fill.L28 = R28[iii * 12 + 0]
                    fill.L29 = R28[iii * 12 + 1]
                    fill.L30 = R28[iii * 12 + 2]
                    fill.K27 = 'R28'
                    fill.K28 = R28[iii * 12 + 3]
                    fill.K29 = R28[iii * 12 + 4]
                    fill.K30 = R28[iii * 12 + 5]
                    fill.J27 = 'R28'
                    fill.J28 = R28[iii * 12 + 6]
                    fill.J29 = R28[iii * 12 + 7]
                    fill.J30 = R28[iii * 12 + 8]
                    fill.I27 = 'R28'
                    fill.I28 = R28[iii * 12 + 9]
                    fill.I29 = R28[iii * 12 + 10]
                    fill.I30 = R28[iii * 12 + 11]
                    R28len = R28len - 12
                if fill.L7 != -1:
                    curSheet.unmerge_cells(start_row=6, end_row=6,
                                           start_column=12, end_column=13)
                    curSheet.unmerge_cells(start_row=7, end_row=7,
                                           start_column=12, end_column=13)
                    # 设置字体格式
                    from openpyxl.styles import Color, Font, Alignment
                    font = Font(u'宋体', size=12, color='000000')
                    curSheet['L6'].font = font
                    curSheet['L7'].font = font
                    curSheet['L6'].alignment = Alignment(horizontal='center', vertical='center')
                    curSheet['L7'].alignment = Alignment(horizontal='center', vertical='center')
                    curSheet['M6'].font = font
                    curSheet['M7'].font = font
                    curSheet['M6'].alignment = Alignment(horizontal='center', vertical='center')
                    curSheet['M7'].alignment = Alignment(horizontal='center', vertical='center')
                    curSheet['L6'] = '膨胀剂'
                    curSheet['L7'] = fill.L7
                    curSheet['M6'] = '外加剂'
                    curSheet['M7'] = fill.M7
                else:
                    curSheet['L7'] = fill.M7
                curSheet['B4'] = fill.B4
                curSheet['G4'] = fill.G4
                curSheet['B6'] = fill.B6
                curSheet['B7'] = fill.B7
                # curSheet['C7'] = fill.C7
                curSheet['G7'] = fill.G7
                curSheet['H7'] = fill.H7
                curSheet['I7'] = fill.I7
                curSheet['J7'] = fill.J7
                curSheet['K7'] = fill.K7
                curSheet['G8'] = fill.G8
                curSheet['K8'] = fill.K8
                curSheet['K9'] = fill.K9
                curSheet['K10'] = fill.K10
                curSheet['C28'] = fill.C28
                curSheet['I27'] = fill.I27
                curSheet['I28'] = fill.I28
                curSheet['I29'] = fill.I29
                curSheet['I30'] = fill.I30
                curSheet['J27'] = fill.J27
                curSheet['J28'] = fill.J28
                curSheet['J29'] = fill.J29
                curSheet['J30'] = fill.J30
                curSheet['K27'] = fill.K27
                curSheet['K28'] = fill.K28
                curSheet['K29'] = fill.K29
                curSheet['K30'] = fill.K30
                curSheet['L28'] = fill.L28
                curSheet['L29'] = fill.L29
                curSheet['L30'] = fill.L30
                Project9Manager = parm.Project9Manager
                Project9Checker = parm.Project9Checker
                Project9Record = parm.Project9Record
                curSheet = insertpic(curSheet, Project9Manager, 'B32')
                curSheet = insertpic(curSheet, Project9Checker, 'D32')
                curSheet = insertpic(curSheet, Project9Record, 'G32', width=60)

    modebook.remove(modebook.worksheets[0])
    return modebook
if __name__ == '__main__':
    filename = '../../工地混凝土使用记录5.xlsx'
    book1 = load_workbook('试件强度实验结果汇总表3.xlsx')
    book2 = load_workbook('配合比设计报告表7.xlsx')
    book = load_workbook(filename)
    ConQualityRecord(book, book1, book2).save('混凝土搅拌质量记录表9.xlsx')

