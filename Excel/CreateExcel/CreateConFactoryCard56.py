# -*- coding:utf-8 -*-
"""
@author: Darcy
@time: 2018/10/30 13:08
"""

from openpyxl import load_workbook
import Bean.ExcelBean
import Bean.DbBean
from Util.ExcelInitUtil import *
import Util.FileUitl
import Util.ExcelUtil
from DateBase.connect_db import session
from Util.query_database import query_mix

def CreateConFacyoryCard(useBook, buyBook):
    """
        获取使用表数据
    """
    useSheets = useBook.sheetnames
    modeName = "../../Mode/5.xlsx"
    modeBook = load_workbook(modeName)
    conUseRecordList = []
    for sheetNum in range(len(useSheets)):
        useSheet = useBook.worksheets[sheetNum]
        rows = useSheet.max_row
        ProjectName = useSheet['A2'].value.replace('工程名称：', '').strip()
        BuildUnit = useSheet['A3'].value.replace('建设单位：', '').strip()
        ConstrUnit = useSheet['A4'].value.replace('施工单位：', '').strip()
        ContractId = useSheet['A5'].value.replace('合同编号：', '').strip()
        ProjectManager = useSheet['A6'].value.replace('项目经理：', '').strip()
        conUseDataList = []
        for rowNum in range(10, rows + 1):
            ProjectSite = useSheet.cell(row=rowNum, column=1).value
            ConcreteName = useSheet.cell(row=rowNum, column=2).value
            ConcreteStrength = useSheet.cell(row=rowNum, column=3).value
            ImperLevel = useSheet.cell(row=rowNum, column=4).value
            SwellLevel = useSheet.cell(row=rowNum, column=5).value
            CuringDate = useSheet.cell(row=rowNum, column=6).value
            StrengthString = ConcreteStrength
            if (ImperLevel == None):
                StrengthString = StrengthString + ConcreteName.replace('砼','')
            else:
                StrengthString = StrengthString + ConcreteName.replace('砼','')+ ImperLevel
            if (SwellLevel != None):
                StrengthString = StrengthString + '\n(' + str(int(SwellLevel * 100)) + '%膨胀)'
            conUseRecord = Bean.ExcelBean.FactoryConUseBean(ProjectName=ProjectName,
                                                            BuildUnit=BuildUnit,
                                                            ConstrUnit=ConstrUnit,
                                                            ContractId=ContractId,
                                                            ProjectManager=ProjectManager,
                                                            ProjectSite=ProjectSite,
                                                            ConcreteName=ConcreteName,
                                                            ImperLevel=ImperLevel,
                                                            SwellLevel=SwellLevel,
                                                            ConcreteStrength=ConcreteStrength,
                                                            CuringDate=CuringDate,
                                                            StrengthString=StrengthString)
            conUseDataList.append(conUseRecord)
            """
                复制模板
            """
            modeSheet = modeBook.worksheets[0]
            # iniexcel5(modeSheet)
            modeSheet = modeBook.copy_worksheet(modeSheet)
            # print(ProjectSite, rowNum)
            iniexcel5(modeSheet)
            modeSheet.title = str(str(sheetNum + 1) + '-' + str(rowNum - 9))
        # conUseDataList.sort(key=lambda x: x.CuringDate, reverse=False)
        conUseRecordList.append(conUseDataList)

    modeBook.remove(modeBook.worksheets[0])
    """
        获取水泥编号
    """
    curNum = 0
    cenBuyList = []
    for sheetNum in range(len(useSheets)):
        sheet = useBook.worksheets[sheetNum]
        useRecordNum = len(conUseRecordList[sheetNum])
        moreFixNum, rows = divmod(useRecordNum, 15)
        if(rows == 0):
            moreFixNum = moreFixNum - 1
        cenDataList = []
        for insertNum in range(moreFixNum + 1):
            buySheet = buyBook.worksheets[curNum]
            InsertFlag = 0
            curNum = curNum + 1
            for buyNum in range(15):
                # print(len(conUseRecordList[sheetNum]), buyNum + 15 * insertNum)
                if((buyNum + 15 * insertNum) == len(conUseRecordList[sheetNum])):
                    break
                CementId = buySheet['F' + str(7 + InsertFlag)].value
                cenDataList.append(CementId)
                InsertFlag = InsertFlag + 1
        cenBuyList.append(cenDataList)
    """
        获取配合比数据
    """
    conMixList = []
    for conUseRecord in conUseRecordList:
        conMixDataList = []
        for conUseData in conUseRecord:
            queryStreng = int(conUseData.ConcreteStrength.strip('C'))
            queryName = conUseData.ConcreteName
            queryImperLevel = conUseData.ImperLevel
            querySwellLevel = conUseData.SwellLevel
            UseMix = query_mix(ConcreteName=queryName, StrengthLevel=queryStreng, ImperLevel=queryImperLevel,
                               SwellLevel=querySwellLevel)
            conMixDataList.append(UseMix)
        conMixList.append(conMixDataList)
    """
        插入数据
    """
    curSheetNum = 0
    for sheetNum in range(len(useSheets)):
        useSheet = useBook.worksheets[sheetNum]
        rows = useSheet.max_row
        for rowNum in range(10, rows + 1):
            resultSheet = modeBook.worksheets[curSheetNum]
            curSheetNum = curSheetNum + 1
            rowIndex = rowNum - 10
            resultSheet['A4'] = "出厂日期：" + conUseRecordList[sheetNum][rowIndex].CuringDate.strftime('%Y.%m.%d')
            resultSheet['e4'] = "工程名称：" + conUseRecordList[sheetNum][rowIndex].ProjectName
            resultSheet['A5'] = "合同编号：" + conUseRecordList[sheetNum][rowIndex].ContractId
            resultSheet['E5'] = "工程部位：" + conUseRecordList[sheetNum][rowIndex].ProjectSite
            resultSheet['A6'] = "建设单位：" + conUseRecordList[sheetNum][rowIndex].BuildUnit
            resultSheet['E6'] = "施工单位：" + conUseRecordList[sheetNum][rowIndex].ConstrUnit
            resultSheet['A7'] = "强度等级：" + conUseRecordList[sheetNum][rowIndex].StrengthString
            resultSheet['E7'] = "工程负责人：" + conUseRecordList[sheetNum][rowIndex].ProjectManager
            resultSheet['C10'] = str(cenBuyList[sheetNum][rowIndex])
            resultSheet['A14'] = '{0:.0f}'.format(float(conMixList[sheetNum][rowIndex].MassDensity)) + "±40" + "kg/m³"
            resultSheet['C14'] = conMixList[sheetNum][rowIndex].SlumpNum
            resultSheet['D14'] = conMixList[sheetNum][rowIndex].InitialTime
            resultSheet['F14'] = conMixList[sheetNum][rowIndex].FinalTime
            # resultSheet['E18'] = "三水区建友混凝土有限公司试验室"
            resultSheet['F20'] = conUseRecordList[sheetNum][rowIndex].CuringDate.strftime('%Y$%#m-%#d').replace('$','年').replace('-', '月') + '日'
            # 3.28
            if(conUseRecordList[sheetNum][rowIndex].SwellLevel != None):
                resultSheet.unmerge_cells(start_row=8, end_row=8,
                                          start_column=7, end_column=8)
                resultSheet.unmerge_cells(start_row=9, end_row=10,
                                          start_column=7, end_column=8)
                resultSheet.merge_cells(start_row=9, end_row=10,
                                          start_column=7, end_column=7)
                resultSheet.merge_cells(start_row=9, end_row=10,
                                        start_column=8, end_column=8)
                resultSheet.unmerge_cells(start_row=11, end_row=11,
                                          start_column=7, end_column=8)
                # 设置字体格式
                from openpyxl.styles import Color, Font, Alignment, Border
                font = Font(u'宋体', size=12, color='000000')
                resultSheet['H8'].font = font
                resultSheet['H9'].font = font
                resultSheet['G9'].font = font
                resultSheet['H8'].alignment = Alignment(horizontal='center', vertical='center')
                resultSheet['H9'].alignment = Alignment(horizontal='center', vertical='center', wrap_text='True')
                resultSheet['G9'].alignment = Alignment(horizontal='center', vertical='center', wrap_text='True')
                resultSheet['H11'].font = font
                resultSheet['H11'].alignment = Alignment(horizontal='center', vertical='center')
                resultSheet['G9'] = "HEA型高效抗裂膨胀剂"
                resultSheet['G8'] = "膨胀剂"
                resultSheet['H8'] = "外加剂"
                resultSheet['H9'] = "LS-JS(B)高效减水剂"
                resultSheet['H11'] = "合格"
            from Util.InsetPicUtil import insertpic
            from Util.get_parm import get_parm
            parm = get_parm()
            resultSheet = insertpic(resultSheet, picname=parm.Project5Manager, position='B17', width=90, heigh=30)
            resultSheet = insertpic(resultSheet, picname=parm.Project5Filler, position='F17', width=90, heigh=30)
            resultSheet = iniexcel5(resultSheet)

    # """
    # 水下砼合并
    # """
    # length = len(modeBook.sheetnames)
    # for i in range(length-1, -1, -1):
    #     if str(modeBook.worksheets[i]['A7'].value)[-2:] == '水下':
    #         for j in range(i - 1, -1, -1):
    #             if str(modeBook.worksheets[i]['A4'].value) == str(modeBook.worksheets[j]['A4'].value):
    #                 modeBook.worksheets[j]['E5'] = str(modeBook.worksheets[j]['E5'].value) + '、' +\
    #                                                str(modeBook.worksheets[i]['E5'].value).split('：')[-1]
    #                 modeBook.remove(modeBook.worksheets[i])
    #                 break

    return modeBook

"""
    生成混凝土坍落度验收表
    Parm:
        CardBook 工地混凝土出厂合格证表
        session: 数据库池
    Return:
        modeBook 生成工作簿
"""
def CreateSlumpCard(cardBook):
    # modeName = "../Mode/6.xlsx"
    modeName = "../../Mode/6.xlsx"
    from Util.get_parm import get_parm
    parm = get_parm()
    modeBook = Util.FileUitl.openFileXlsx(modeName)
    modeBook = Util.ExcelUtil.CopyFromModeExcel(modeBook, cardBook)
    modeBook[modeBook.sheetnames[0]].title = "1-1"
    cardSheets = cardBook.sheetnames
    for sheetNum in range(len(cardSheets)):
        """
            获得出厂合格证数据
        """
        cardSheet = cardBook.worksheets[sheetNum]
        ProjectName = cardSheet['E4'].value.strip('工程名称：')
        ConstrUnit = cardSheet['E6'].value.strip('施工单位：')
        ProjectManager = cardSheet['E7'].value.strip('工程负责人：')
        ConcreteStrength = cardSheet['A7'].value.strip('强度等级：')
        CuringDate = cardSheet['A4'].value.strip('出厂日期：')
        ProjectSite = cardSheet['E5'].value.strip('工程部位：')
        SlumpNum = cardSheet['C14'].value
        """
            插入数据
        """
        resultSheet = modeBook.worksheets[sheetNum]
        resultSheet['A6'] = "工程名称：" + ProjectName
        resultSheet['A7'] = "施工单位：" + ConstrUnit
        resultSheet['E6'] = "浇筑部位：" + ProjectSite
        resultSheet['A8'] = "工程负责人：" + ProjectManager
        resultSheet['E7'] = "强度等级：" + ConcreteStrength
        resultSheet['A5'] = "供应时间：" + CuringDate
        resultSheet['E8'] = "工地要求坍落度：" + SlumpNum
        resultSheet = iniexcel6(resultSheet)
    return modeBook

if __name__ == '__main__':
    session = session
    filename = "..\..\工地混凝土使用记录5.xlsx"
    useBook = load_workbook(filename)
    useSheets = useBook.sheetnames
    modeName = "../../Mode/5.xlsx"
    modeBook = load_workbook(modeName)
    buyPath = r"水泥购进，使用一览表1.xlsx"
    buyBook = load_workbook(buyPath)
    cardBook = CreateConFacyoryCard(useBook,buyBook)
    cardBook.save("混凝土出厂合格证5.xlsx")
    SlumpBook = CreateSlumpCard(cardBook)
    SlumpBook.save("混凝土坍落度验收表6.xlsx")
