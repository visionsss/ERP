# coding=utf-8
"""
Author: vision
date: 2019/4/7 15:55
"""
from openpyxl import load_workbook
from Util.ExcelInitUtil import iniexcel
from DateBase.creat import ConcreteUseRecord
from Util.get_parm import get_parm
from Util.query_database import query_mix
from datetime import timedelta
from random import randint
from Util.InsetPicUtil import insertpic

def fun(list1):
    set1 = set(list1)
    for i in set1:
        if list1.count(i) == 1:
            continue
        else:
            count = 0
            for j in range(len(list1)):
                if list1[j] == i:
                    count += 1
                    if count < 10:
                        list1[j] = str(list1[j]) + '-0' + str(count)
                    else:
                        list1[j] = str(list1[j]) + '-' + str(count)

    return list1


def findWaterAsh(
        i,
        ConcreteName,
        ConcreteStrength,
        ImperLevel,
        SwellLevel,
        book2):
    if ImperLevel is None:
        ImperLevel = '/'
    if SwellLevel is None:
        SwellLevel = '/'
    else:
        SwellLevel *= 100
        SwellLevel = str(int(SwellLevel)) + '%'
    # print(i, ConcreteName, ConcreteStrength, ImperLevel, SwellLevel)
    for ii in book2.sheetnames:
        if (str(book2[ii].cell(row=9, column=2).value) == str(ConcreteName) and
            str(book2[ii].cell(row=9, column=8).value) == str(ConcreteStrength) and
            str(book2[ii].cell(row=9, column=12).value) == str(ImperLevel) and
                str(book2[ii].cell(row=9, column=13).value) == str(SwellLevel)):
                # print(ii)
            WaterNum = book2[ii].cell(row=25, column=2).value
            Modu = book2[ii].cell(row=13, column=12).value
            return WaterNum, Modu


class Sheet:
    def __init__(self):
        self.E3 = ''
        self.A4 = ''
        self.A5 = ''
        self.A6 = ''
        self.D6 = ''
        self.F6 = ''
        self.D7 = ''
        self.D8 = ''
        self.D9 = ''
        self.D10 = ''
        self.D11 = ''
        self.D13 = ''
        self.D14 = ''
        self.D18 = ''
        self.D25 = ''
        self.D27 = ''


'''
 useBook,使用记录表
 book2,配合比设计报告表
 session
 10 抗渗
'''


def PermeabilityTestReport(useBook, book2):
    modeName = "../../Mode/10.xlsx"
    modebook = load_workbook(modeName)
    useSheets = useBook.sheetnames
    UseRecordListArr = []
    """
        获取数据
    """
    ProjectNames = []
    for sheetNum in range(len(useSheets)):
        sheet = useBook.worksheets[sheetNum]
        ProjectNames.append(sheet.cell(row=2, column=1).value)
        UseRecordList = []
        rows = sheet.max_row
        for rowNum in range(10, rows + 1):
            if sheet.cell(row=rowNum, column=1).value is None:
                break
            ProjectSite = sheet.cell(row=rowNum, column=1).value
            ConcreteName = sheet.cell(row=rowNum, column=2).value
            ConcreteStrength = sheet.cell(row=rowNum, column=3).value
            ImperLevel = sheet.cell(row=rowNum, column=4).value
            SwellLevel = sheet.cell(row=rowNum, column=5).value
            CuringDate = sheet.cell(row=rowNum, column=6).value
            CuringTime = sheet.cell(row=rowNum, column=7).value
            CuringNum = sheet.cell(row=rowNum, column=8).value
            ConUseRecord = ConcreteUseRecord(
                ProjectSite=ProjectSite,
                ConcreteName=ConcreteName,
                ConcreteStrength=ConcreteStrength,
                ImperLevel=ImperLevel,
                SwellLevel=SwellLevel,
                CuringDate=CuringDate,
                CuringTime=CuringTime,
                CuringNum=CuringNum)
            UseRecordList.append(ConUseRecord)
        UseRecordListArr.append(UseRecordList)
    '''
        填数据
    '''
    parm = get_parm()
    Min = float(parm.Project10MinCreepEdit)
    Max = float(parm.Project10MaxCreepEdit)
    for sheetNum in range(len(useSheets)):
        useSheet = useBook.worksheets[sheetNum]
        rows = useSheet.max_row
        ProjectName = ProjectNames[sheetNum]
        count = 0
        num = 0
        for rowNum in range(10, rows + 1):
            useMessage = UseRecordListArr[sheetNum][count]
            count += 1
            if useMessage.ImperLevel is not None:
                num += 1
                curSheet = modebook.copy_worksheet(modebook.worksheets[0])
                curSheet.title = (str(sheetNum + 1) + '#' + str(num))
                iniexcel(curSheet, 10)
                ConcreteName = str(useMessage.ConcreteName)
                ConcreteStrength = str(useMessage.ConcreteStrength).replace('C', '')
                if useMessage.ImperLevel != None:
                    ImperLevelx = str(useMessage.ImperLevel)
                else:
                    ImperLevelx = None
                if useMessage.SwellLevel != None:
                    SwellLevelx = str(useMessage.SwellLevel)
                else:
                    SwellLevelx = None
                UseMix = query_mix(ConcreteName,ConcreteStrength,ImperLevelx,SwellLevelx)
                # UseMix = session.query(
                #     Bean.DbBean.ConcreteMix).filter(
                #     Bean.DbBean.ConcreteMix.ConcreteName == useMessage.ConcreteName,
                #     Bean.DbBean.ConcreteMix.StrengthLevel == useMessage.ConcreteStrength.replace(
                #         'C',
                #         ''),
                #     Bean.DbBean.ConcreteMix.ImperLevel == useMessage.ImperLevel,
                #     Bean.DbBean.ConcreteMix.SwellLevel == useMessage.SwellLevel).first()
                # print(useMessage.ConcreteName,useMessage.ConcreteStrength,useMessage.ImperLevel,useMessage.SwellLevel)
                # print(UseMix)
                fill = Sheet()
                fill.E3 = parm.Project10InspectCodeEdit
                fill.A4 = ProjectName
                fill.A5 = '工程部位：' + useMessage.ProjectSite
                fill.A6 = '检验日期：' + \
                    str((useMessage.CuringDate + timedelta(days=+28)).strftime('%Y-%#m-%#d'))
                fill.D6 = '报告日期：' + \
                    str((useMessage.CuringDate + timedelta(days=+31)).strftime('%Y-%#m-%#d'))
                fill.F6 = '检验依据：' + \
                    str(parm.Project10ConTestReportTestBasisEdit)
                fill.D9 = 'SJY' + \
                    str(useMessage.CuringDate).replace('00:00:00', '').replace('-', '')
                fill.D7 = useMessage.ConcreteStrength + \
                    useMessage.ConcreteName.replace('砼', '')
                # print(useMessage.ConcreteStrength,' ', useMessage.ConcreteName)
                # print(useMessage.ConcreteStrength + useMessage.ConcreteName.replace('砼', ''))
                fill.D8 = useMessage.ImperLevel
                fill.D10 = str(useMessage.CuringDate.strftime('%Y-%#m-%#d'))
                if useMessage.ImperLevel == 'P6':
                    fill.D11 = str((useMessage.CuringDate + timedelta(days=+28)).strftime('%Y-%#m-%#d')) + '  ' + str(parm.Project10TimeEdit) + '  -   '\
                        + str((useMessage.CuringDate + timedelta(days=+30)).strftime('%Y-%#m-%#d')) + '  ' + str(parm.Project10TimeEdit)
                if useMessage.ImperLevel == 'P8':
                    s = int(parm.Project10TimeEdit.split(':')[0])
                    m = parm.Project10TimeEdit.split(':')[1]
                    if s >= 8:
                        fill.D11 = str(useMessage.CuringDate + timedelta(days=+28)).replace('00:00:00', '') + str(parm.Project10TimeEdit) + \
                            '  -   ' + str(useMessage.CuringDate + timedelta(days=+31)).replace('00:00:00', '') + str(s - 8) + ':' + str(m)
                    else:
                        fill.D11 = str(useMessage.CuringDate + timedelta(days=+28)).replace('00:00:00', '') + str(parm.Project10TimeEdit) + '  -   '\
                            + str(useMessage.CuringDate + timedelta(days=+31)).replace('00:00:00', '') + str(s + 16) + ':' + str(m)
                fill.D13 = findWaterAsh(
                    sheetNum + 1,
                    useMessage.ConcreteName,
                    useMessage.ConcreteStrength,
                    useMessage.ImperLevel,
                    useMessage.SwellLevel,
                    book2)[0]
                # print(fill.D13)
                # fill.D18 = fill.D13 = findWaterAsh(sheetNum+1, useMessage.ConcreteName, useMessage.ConcreteStrength,
                #                         useMessage.ImperLevel, useMessage.SwellLevel, book2)[1]
                # fill.D13 = '0.28'
                fill.D18 = '2.5'
                fill.D14 = UseMix.MixRatioName
                fill.D25 = UseMix.AdmixtureAmount
                fill.D27 = '以下是本组（六块）的最高渗水值(mm)：' + '\n1#' +\
                           str(randint(Min, Max)) + '；2#' +\
                           str(randint(Min, Max)) + '；3#' +\
                           str(randint(Min, Max)) + '；4#' +\
                           str(randint(Min, Max)) + '；5#' +\
                           str(randint(Min, Max)) + '；6#' +\
                           str(randint(Min, Max)) + '。'
                curSheet['E3'] = fill.E3
                curSheet['A4'] = fill.A4
                curSheet['A5'] = fill.A5
                curSheet['A6'] = fill.A6
                curSheet['D6'] = fill.D6
                curSheet['F6'] = fill.F6
                curSheet['D7'] = fill.D7
                curSheet['D8'] = fill.D8
                curSheet['D9'] = fill.D9
                curSheet['D10'] = fill.D10
                curSheet['D11'] = fill.D11
                curSheet['D13'] = fill.D13
                curSheet['D14'] = fill.D14
                curSheet['D18'] = fill.D18
                curSheet['D25'] = fill.D25
                curSheet['D27'] = fill.D27
                curSheet = insertpic(curSheet,parm.Project10Manager,'C29',width=80, heigh=30)
                curSheet = insertpic(curSheet, parm.Project10Examine, 'D29')
                curSheet = insertpic(curSheet, parm.Project10Checker, 'F29')
    Idlist = []
    for i in range(len(modebook.worksheets)):
        Idlist.append(modebook.worksheets[i].cell(row=9, column=4).value)
    # print(Idlist)
    Idlist = fun(Idlist)
    for i in range(len(modebook.worksheets)):
        modebook.worksheets[i]['D9'] = Idlist[i]
    if len(modebook.sheetnames) > 1:
        modebook.remove(modebook.worksheets[0])
    else:
        iniexcel(modebook.worksheets[0], 10)
    return modebook


if __name__ == '__main__':
    usefile = '../../工地混凝土使用记录test.xlsx'
    uesbook = load_workbook(usefile)
    book2 = load_workbook('配合比设计报告表7.xlsx')
    PermeabilityTestReport(uesbook, book2).save('抗渗性能检测报告10.xlsx')
