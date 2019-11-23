# coding=utf-8
"""
Author: vision
date: 2019/4/7 18:08
"""
from openpyxl import load_workbook
from Util.ExcelInitUtil import iniexcel
from DateBase.creat import ConcreteUseRecord
from Util.get_parm import get_parm
from Util.query_database import query_mix
from datetime import timedelta
from random import randint
from Util.InsetPicUtil import insertpic


'''
 useBook,使用记录表
 session
 10 抗渗
'''


def PermeabilityTestReportMode(useBook):
    modeName = "../../Mode/10.xlsx"
    modebook = load_workbook(modeName)
    useSheets = useBook.sheetnames
    UseRecordListArr = []
    """
        获取数据
    """
    ProjectNames = []
    for sheetNum in range(len(useSheets)):
        sheet = useBook.worksheets[sheetNum]
        ProjectNames.append(sheet.cell(row=2, column=1).value)
        UseRecordList = []
        rows = sheet.max_row
        for rowNum in range(10, rows + 1):
            if sheet.cell(row=rowNum, column=1).value is None:
                break
            ProjectSite = sheet.cell(row=rowNum, column=1).value
            ConcreteName = sheet.cell(row=rowNum, column=2).value
            ConcreteStrength = sheet.cell(row=rowNum, column=3).value
            ImperLevel = sheet.cell(row=rowNum, column=4).value
            SwellLevel = sheet.cell(row=rowNum, column=5).value
            CuringDate = sheet.cell(row=rowNum, column=6).value
            CuringTime = sheet.cell(row=rowNum, column=7).value
            CuringNum = sheet.cell(row=rowNum, column=8).value
            ConUseRecord = ConcreteUseRecord(
                ProjectSite=ProjectSite,
                ConcreteName=ConcreteName,
                ConcreteStrength=ConcreteStrength,
                ImperLevel=ImperLevel,
                SwellLevel=SwellLevel,
                CuringDate=CuringDate,
                CuringTime=CuringTime,
                CuringNum=CuringNum)
            UseRecordList.append(ConUseRecord)
        UseRecordListArr.append(UseRecordList)
    '''
        填数据
    '''
    parm = get_parm()
    Min = float(parm.Project10MinCreepEdit)
    Max = float(parm.Project10MaxCreepEdit)
    for sheetNum in range(len(useSheets)):
        useSheet = useBook.worksheets[sheetNum]
        rows = useSheet.max_row
        ProjectName = ProjectNames[sheetNum]
        count = 0
        num = 0
        for rowNum in range(10, rows + 1):
            useMessage = UseRecordListArr[sheetNum][count]
            count += 1
            if useMessage.ImperLevel is not None:
                num += 1
                curSheet = modebook.copy_worksheet(modebook.worksheets[0])
                curSheet.title = (str(sheetNum + 1) + '#' + str(num))
                iniexcel(curSheet, 10)

    Idlist = []
    for i in range(len(modebook.worksheets)):
        Idlist.append(modebook.worksheets[i].cell(row=9, column=4).value)
    # print(Idlist)
    for i in range(len(modebook.worksheets)):
        modebook.worksheets[i]['D9'] = Idlist[i]
    if len(modebook.sheetnames) > 1:
        modebook.remove(modebook.worksheets[0])
    else:
        iniexcel(modebook.worksheets[0], 10)
    return modebook


if __name__ == "__main__":
    use_file = '../../工地混凝土使用记录.xlsx'
    use_book = load_workbook(use_file)
    PermeabilityTestReportMode(use_book).save('10voidMode.xlsx')
