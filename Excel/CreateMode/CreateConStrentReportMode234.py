# coding=utf-8
"""
Author: vision
date: 2019/4/8 9:56
"""
from openpyxl import load_workbook
from Util.ExcelInitUtil import iniexcel2
from Util.ExcelInitUtil import iniexcel3
from Util.ExcelInitUtil import iniexcel4

"""
    生成混凝土试件抗压强度检验报告,混凝土试件强度实验结果汇总表
    Parm:
        useBook 使用记录表
        session 数据库连接池
        count 使用记录数量
    Return:
        resultRepoBook      混凝土试件抗压强度检验报告
        resultRepoSumBook   混凝土试件强度实验结果汇总表
        resultRepoCompBook  标准养护混凝土抗压强度计算表
"""
def ConStrengReportMode(useBook):
    secModeName = "../../Mode/2.xlsx"
    thrModeNmae = "../../Mode/3.xlsx"
    fouModeName = "../../Mode/4.xlsx"
    useSheets = useBook.sheetnames
    resultRepoBook = load_workbook(secModeName)
    resultRepoSumBook = load_workbook(thrModeNmae)
    resultRepoCompBook = load_workbook(fouModeName)

    """
           复制模板
       """
    for sheetNum in range(len(useSheets)):
        modeSheet = resultRepoBook.worksheets[0]
        iniexcel2(modeSheet)
        modeSheet = resultRepoBook.copy_worksheet(modeSheet)
        iniexcel2(modeSheet)
        modeSheet.title = str(str(sheetNum + 1))
    resultRepoBook.remove(resultRepoBook.worksheets[0])

    for sheetNum in range(len(useSheets)):
        modeSheet = resultRepoSumBook.worksheets[0]
        iniexcel3(modeSheet)
        modeSheet = resultRepoSumBook.copy_worksheet(modeSheet)
        iniexcel3(modeSheet)
        modeSheet.title = str(str(sheetNum + 1))
    resultRepoSumBook.remove(resultRepoSumBook.worksheets[0])

    for sheetNum in range(len(useSheets)):
        modeSheet = resultRepoCompBook.worksheets[0]
        iniexcel4(modeSheet)
        modeSheet = resultRepoCompBook.copy_worksheet(modeSheet)
        iniexcel4(modeSheet)
        modeSheet.title = str(sheetNum + 1)
    resultRepoCompBook.remove(resultRepoCompBook.worksheets[0])
    return resultRepoBook, resultRepoSumBook, resultRepoCompBook

if __name__ == "__main__":
    filename = "..\..\工地混凝土使用记录.xlsx"
    useBook = load_workbook(filename)
    resultRepoBook, resultRepoSumBook, resultRepoCompBook = ConStrengReportMode(useBook)
    resultRepoBook.save('标准养护混凝土试件抗压强度检验报告2.xlsx')
    resultRepoSumBook.save("试件强度实验结果汇总表3.xlsx")
    resultRepoCompBook.save('标准养护混凝土抗压强度计算表4.xlsx')
    pass
