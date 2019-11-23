# coding=utf-8
"""
Author: vision
date: 2019/4/8 9:50
"""
from openpyxl import load_workbook
from Util.ExcelInitUtil import iniexcel1, iniexcel7
import Bean.ExcelBean
import Bean.DbBean
import Util.ExcelUtil
from collections import defaultdict
from datetime import timedelta
from Util.get_parm import get_parm

"""
    生成水泥购进，使用一览表
    Parm:
        useBook 使用记录表
        session 数据库连接池
    Return:
        modeBook 购进使用表
        CementDesignBook 配合比设计报告表
        maxRows 使用记录数量
"""
def ConUseBuyRecordMode(useBook):
    useSheets = useBook.sheetnames
    modeName = "../../Mode/1.xlsx"
    modeBook = load_workbook(modeName)
    """
        第一份表操作
    """
    """
        复制模板
    """
    for sheetNum in range(len(useSheets)):
        useSheet = useBook.worksheets[sheetNum]
        rows = useSheet.max_row - 9
        moreFixNum, rows = divmod(rows, 15)
        if rows == 0:
            moreFixNum = moreFixNum - 1
        for copyNum in range(moreFixNum + 1):
            modeSheet = modeBook.worksheets[0]
            iniexcel1(modeSheet)
            modeSheet = modeBook.copy_worksheet(modeSheet)
            iniexcel1(modeSheet)
            modeSheet.title = str(str(sheetNum + 1) + '-' + str(copyNum + 1))

    modeName = "../../Mode/7.xlsx"
    CementDesignBook = load_workbook(modeName)

    """
        复制模板
    """
    for sheetNum in range(len(useSheets)):
        modeSheet = CementDesignBook.worksheets[0]
        iniexcel7(modeSheet)
        modeSheet = CementDesignBook.copy_worksheet(modeSheet)
        iniexcel7(modeSheet)
        modeSheet.title = str(str(sheetNum + 1))
    CementDesignBook.remove(CementDesignBook.worksheets[0])
    modeBook.remove(modeBook.worksheets[0])

    return modeBook, CementDesignBook

if __name__ == "__main__":
    filename = "..\..\工地混凝土使用记录.xlsx"
    useBook = load_workbook(filename)
    result, MixDesignBook = ConUseBuyRecordMode(useBook)
    result.save('水泥购进，使用一览表1.xlsx')
    MixDesignBook.save("配合比设计报告表7.xlsx")
    pass
