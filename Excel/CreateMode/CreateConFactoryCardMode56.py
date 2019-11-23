# coding=utf-8
"""
Author: vision
date: 2019/4/8 10:04
"""
from openpyxl import load_workbook
from Util.ExcelInitUtil import *

def CreateConFacyoryCardMode(useBook):
    useSheets = useBook.sheetnames
    modeName = "../../Mode/5.xlsx"
    modeBook = load_workbook(modeName)
    for sheetNum in range(len(useSheets)):
        modeSheet = modeBook.worksheets[0]
        iniexcel5(modeSheet)
        modeSheet = modeBook.copy_worksheet(modeSheet)
        iniexcel5(modeSheet)
        modeSheet.title = str(sheetNum + 1)
    modeBook.remove(modeBook.worksheets[0])
    return modeBook

def CreateSlumpCardMode(useBook):
    useSheets = useBook.sheetnames
    modeName = "../../Mode/6.xlsx"
    modeBook = load_workbook(modeName)
    for sheetNum in range(len(useSheets)):
        modeSheet = modeBook.worksheets[0]
        iniexcel6(modeSheet)
        modeSheet = modeBook.copy_worksheet(modeSheet)
        iniexcel6(modeSheet)
        modeSheet.title = str(sheetNum + 1)
    modeBook.remove(modeBook.worksheets[0])
    return modeBook

if __name__ == "__main__":
    if __name__ == '__main__':
        filename = "..\..\工地混凝土使用记录.xlsx"
        useBook = load_workbook(filename)
        useSheets = useBook.sheetnames
        cardBook = CreateConFacyoryCardMode(useBook)
        cardBook.save("混凝土出厂合格证5.xlsx")
        SlumpBook = CreateSlumpCardMode(cardBook)
        SlumpBook.save("混凝土坍落度验收表6.xlsx")
