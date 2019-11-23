# coding=utf-8
"""
Author: vision
date: 2019/4/7 18:31
"""
from openpyxl import load_workbook
from Util.ExcelInitUtil import iniexcel
from DateBase.creat import ConcreteUseRecord


def findR28(i, j, book1):
    i = str(i)
    j = str(j)
    for ii in book1.sheetnames:
        num = ii.split('-')[0]
        if num == str(i):
            for row in range(7, 27):
                if str(book1[ii].cell(row=row, column=1).value) == j:
                    s = book1[ii].cell(row=row, column=6).value
                    s = str(s)
                    s = s.split(r"/")

                    return s
from Excel.CreateExcel.CreateConStrengReport234 import ConStrengReport


def ConQualityRecordMode(useBook):
    book1 = ConStrengReport(useBook)[1]
    modeName = "../../Mode/9.xlsx"
    modebook = load_workbook(modeName)
    useSheets = useBook.sheetnames
    UseRecordListArr = []
    """
        获取数据
    """
    ProjectNames = []
    ProjectUnits = []
    for sheetNum in range(len(useSheets)):
        sheet = useBook.worksheets[sheetNum]
        ProjectNames.append(sheet.cell(row=2, column=1).value)
        ProjectUnits.append(sheet.cell(row=4, column=1).value)
        UseRecordList = []
        rows = sheet.max_row
        for rowNum in range(10, rows + 1):
            if sheet.cell(row=rowNum, column=1).value is None:
                break
            ProjectSite = sheet.cell(row=rowNum, column=1).value
            ConcreteName = sheet.cell(row=rowNum, column=2).value
            ConcreteStrength = sheet.cell(row=rowNum, column=3).value
            ImperLevel = sheet.cell(row=rowNum, column=4).value
            SwellLevel = sheet.cell(row=rowNum, column=5).value
            CuringDate = sheet.cell(row=rowNum, column=6).value
            CuringTime = sheet.cell(row=rowNum, column=7).value
            CuringNum = sheet.cell(row=rowNum, column=8).value
            ConUseRecord = ConcreteUseRecord(
                ProjectSite=ProjectSite,
                ConcreteName=ConcreteName,
                ConcreteStrength=ConcreteStrength,
                ImperLevel=ImperLevel,
                SwellLevel=SwellLevel,
                CuringDate=CuringDate,
                CuringTime=CuringTime,
                CuringNum=CuringNum)
            UseRecordList.append(ConUseRecord)
        UseRecordListArr.append(UseRecordList)
    '''
       填数据
       '''

    curSheetNum = 0
    for sheetNum in range(len(useSheets)):
        useSheet = useBook.worksheets[sheetNum]
        rows = useSheet.max_row
        ProjectName = ProjectNames[sheetNum].replace('工程名称：', '')
        ProjectUnit = ProjectUnits[sheetNum].replace('施工单位：', '')
        count = 0
        for rowNum in range(10, rows + 1):
            useMessage = UseRecordListArr[sheetNum][count]
            count += 1
            R28 = findR28(sheetNum + 1, rowNum - 9, book1)
            for iii in range((len(R28) + 11) // 12):
                curSheet = modebook.copy_worksheet(modebook.worksheets[0])
                curSheet.title = str(sheetNum + 1) + '#' + str(rowNum - 9)
                iniexcel(curSheet, 9)

    modebook.remove(modebook.worksheets[0])
    return modebook


if __name__ == "__main__":
    use_file = '../../工地混凝土使用记录.xlsx'
    use_book = load_workbook(use_file)
    book1 = load_workbook('3、混凝土试块强度试验结果汇总表.xlsx')
    ConQualityRecordMode(use_book).save('9voidMode.xlsx')
